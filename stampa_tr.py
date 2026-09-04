# ============================================================
# MODULO: STAMPA TR
# Foglio "tr" (tabellone giornaliero di rintracciabilita' latte/prodotti).
#
# RISCRITTO DA ZERO IL 29/08 su richiesta esplicita dell'utente: le versioni
# precedenti copiavano e modificavano il template Excel originale (righe
# inserite/eliminate dinamicamente dentro un foglio con celle fuse) - questo
# ha causato piu' crash reali (openpyxl non tiene aggiornate in modo
# affidabile le fusioni quando si inseriscono/eliminano righe piu' volte di
# seguito, un limite noto della libreria, non risolvibile in modo affidabile
# lato nostro). Ora il foglio si COSTRUISCE da zero, riga per riga, in un
# unico passaggio dall'alto in basso: si raccolgono prima TUTTI i dati del
# giorno, si conta quante righe servono, e si scrive tutto in sequenza -
# NESSUN insert_rows/delete_rows, quindi nessuna fusione puo' corrompersi.
# MBC e RBC restano invariati (usano ancora il template originale - li' non
# c'e' mai stato nessun problema, perche' non inseriscono/eliminano righe).
#
# REGOLE (confermate con l'utente nelle sessioni del 27-29/08):
#   - un conferitore ATTIVO compare sempre nella sua categoria (anche a 0 kg
#     quel giorno) - allevamenti/caseifici dop, non-dop, vaccino, cagliata.
#   - i blocchi DOP (allevamenti/caseifici dop) esistono solo se il
#     caseificio stesso e' DOP.
#   - CONGELATO = SOLO gli eventi di SCONGELAMENTO avvenuti quel giorno
#     (latte che rientra in produzione, sempre bufala NON-DOP) - compare
#     solo nei giorni in cui e' successo davvero. Il latte che ESCE per
#     essere congelato va invece nella sezione Lavorazione.
#   - CAGLIATA: righe distinte per bufala e vaccina, ESCLUSE dai totali di
#     latte (sono conferimenti di semilavorato, non di latte).
#   - Acidita'/temperatura: generate automaticamente (come nel vecchio
#     foglio Excel del caseificio, che usava gia' RANDBETWEEN) - confermato
#     dall'utente che devono restare cosi', non diventano campi reali.
#   - Vendite di latte: NESSUN tetto massimo, tutte quelle del giorno.
#   - Prodotti finiti: SOLO quelli realmente prodotti (kg_totale>0) almeno
#     un giorno nel PERIODO selezionato (non nel singolo giorno) - se
#     prodotto anche un solo giorno del periodo, compare in tutti i giorni
#     del periodo (a 0 nei giorni in cui quel giorno specifico non l'ha
#     fatto), per poter confrontare colonna per colonna.
# ============================================================
import random
import datetime as _dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import registro_calc

FOGLIO = "tr"
N_COLONNE = 7

FONT_TITOLO = Font(bold=True, size=14, color="1F3864")
FONT_SOTTOTITOLO_INFO = Font(size=10, italic=True, color="595959")
FONT_SEZIONE = Font(bold=True, size=11, color="FFFFFF")
FILL_SEZIONE = PatternFill("solid", fgColor="1F3864")
FONT_SOTTOSEZIONE = Font(bold=True, size=10, color="1F3864")
FONT_INTESTAZIONE = Font(bold=True, size=9, color="1F3864")
FILL_INTESTAZIONE = PatternFill("solid", fgColor="D9E1F2")
FONT_TOTALE = Font(bold=True, size=10)
FILL_TOTALE = PatternFill("solid", fgColor="F2F2F2")
FONT_VUOTO = Font(italic=True, size=9, color="A6A6A6")
ALLINEA_CENTRO = Alignment(horizontal="center", vertical="center")
BORDO_INTESTAZIONE = Border(bottom=Side(style="thin", color="9DB2D9"))


def _acidita_random():
    return f"3,{random.randint(5, 9)} °SH/50ml"


def _temperatura_random():
    return f"T°C {random.randint(3, 5)},{random.randint(1, 9)}"


# ------------------------------------------------------------
# BLOCCO: RACCOLTA DATI (invariato nella sostanza rispetto alla versione
# precedente - qui cambia solo COME viene scritto sul foglio, non da dove
# arrivano i numeri)
# ------------------------------------------------------------
def _caseificio_is_dop(client, caseificio_id):
    c = client.table("caseifici").select("*").eq("id", caseificio_id).single().execute().data
    return c, bool(c and c.get("is_dop"))


def _conferitori_attivi_categoria(client, caseificio_id, tipi_conferitore, tipo_latte):
    """Tutti i conferitori ATTIVI di quella categoria, indipendentemente da cosa hanno
    conferito quel giorno - determina se il blocco esiste e quali righe mostrare."""
    conferitori = (
        client.table("conferitori")
        .select("*, conferitori_tipi_latte(tipo_latte)")
        .eq("caseificio_id", caseificio_id)
        .eq("attivo", True)
        .in_("tipo", tipi_conferitore)
        .order("ordine")
        .execute()
        .data
    )
    return [
        c for c in conferitori
        if tipo_latte in [t["tipo_latte"] for t in c.get("conferitori_tipi_latte", [])]
    ]


def _righe_dati_categoria(client, conferitori, data_giorno):
    """Una riga per OGNI conferitore attivo, anche con kg=0 quel giorno."""
    righe = []
    for c in conferitori:
        rec = (
            client.table("conferimenti").select("*")
            .eq("conferitore_id", c["id"]).eq("data", str(data_giorno))
            .execute().data
        )
        kg = float(rec[0]["kg"]) if rec and rec[0].get("kg") else 0.0
        ddt = rec[0].get("ddt") if rec else ""
        righe.append({"provenienza": c["ragione_sociale"], "codice": c.get("piva", ""), "kg": kg, "ddt": ddt or ""})
    return righe


def _righe_congelato(client, caseificio_id, data_giorno):
    """SOLO gli eventi di SCONGELAMENTO di quel giorno - una riga per struttura esterna."""
    movimenti = (
        client.table("movimenti_congelato").select("*")
        .eq("caseificio_id", caseificio_id).eq("data", str(data_giorno)).eq("tipo", "scongelamento")
        .execute().data
    )
    per_struttura = {}
    for m in movimenti:
        chiave = m.get("struttura_esterna") or "(interno)"
        d = per_struttura.setdefault(chiave, {"provenienza": chiave, "codice": "", "kg": 0.0, "ddt": []})
        d["kg"] += float(m.get("kg") or 0)
        if m.get("ddt"):
            d["ddt"].append(m["ddt"])
    righe = []
    for d in per_struttura.values():
        d["ddt"] = ", ".join(d["ddt"])
        righe.append(d)
    return righe


def _prodotti_ammessi_nel_periodo(client, caseificio_id, data_da, data_a):
    """Prodotti (attivi, visibili in Produzioni) prodotti REALMENTE (kg_totale>0) almeno
    un giorno nel periodo - un prodotto mai fatto nel periodo non compare in nessun giorno."""
    prodotti = (
        client.table("prodotti").select("*")
        .eq("caseificio_id", caseificio_id).eq("attivo", True).eq("mostra_in_produzioni", True)
        .order("nome").execute().data
    )
    if not prodotti:
        return []
    ids = [p["id"] for p in prodotti]
    produzioni_periodo = (
        client.table("produzioni").select("prodotto_id, kg_totale")
        .in_("prodotto_id", ids).gte("data", str(data_da)).lte("data", str(data_a))
        .execute().data
    )
    ids_usati = {r["prodotto_id"] for r in produzioni_periodo if float(r.get("kg_totale") or 0) > 0}
    return [p for p in prodotti if p["id"] in ids_usati]


def _prodotti_finiti(client, prodotti_ammessi, data_giorno):
    righe = []
    for p in prodotti_ammessi:
        rec = (
            client.table("produzioni").select("*").eq("prodotto_id", p["id"]).eq("data", str(data_giorno))
            .execute().data
        )
        righe.append({
            "nome": p["nome"],
            "totale": float(rec[0]["kg_totale"]) if rec and rec[0].get("kg_totale") else 0.0,
            "diretta": float(rec[0]["kg_diretta"]) if rec and rec[0].get("kg_diretta") else 0.0,
            "terzi": float(rec[0]["kg_terzi"]) if rec and rec[0].get("kg_terzi") else 0.0,
        })
    return righe


def _resa_dop_giorno(client, caseificio_id, data_giorno):
    from stampa_mbc import get_prodotto_e_produzione_giorno
    _, produzione_mozz = get_prodotto_e_produzione_giorno(
        client, caseificio_id, data_giorno, "Mozzarella di Bufala Campana DOP", solo_dop=True
    )
    kg_mozz = float((produzione_mozz or {}).get("kg_totale") or 0)
    kg_trasf_dop = registro_calc.trasformato(client, caseificio_id, "bufala_dop", data_giorno)
    return (kg_mozz / kg_trasf_dop * 100) if kg_trasf_dop > 0 else None


def _latte_dop_per_gruppo(client, caseificio_id, data_giorno, resa, is_dop_prodotto, solo_nome=None):
    """kg di latte DOP consumato per un gruppo di prodotti (dop_altri o declassati) - stessa
    logica del Registro: kg_dop_prodotto / resa."""
    if not resa:
        return 0.0
    prodotti = (
        client.table("prodotti").select("*")
        .eq("caseificio_id", caseificio_id).eq("attivo", True).eq("is_dop", is_dop_prodotto)
        .execute().data
    )
    prodotti = [p for p in prodotti if "ricotta" not in p["nome"].lower()]
    if solo_nome:
        prodotti = [p for p in prodotti if solo_nome.lower() in p["nome"].lower()]
    elif is_dop_prodotto:
        prodotti = [p for p in prodotti if "senza lattosio" not in p["nome"].lower()]

    totale_latte = 0.0
    for p in prodotti:
        rec = (
            client.table("produzioni").select("*, produzione_origine(*)")
            .eq("prodotto_id", p["id"]).eq("data", str(data_giorno)).execute().data
        )
        if not rec:
            continue
        r = rec[0]
        tot = float(r.get("kg_totale") or 0)
        origini = {o["origine"]: o for o in r.get("produzione_origine", [])}
        kg_nondop = float(origini["non_dop"]["kg"]) if "non_dop" in origini and origini["non_dop"].get("kg") else 0.0
        kg_dop = tot - kg_nondop
        if kg_dop > 0:
            totale_latte += kg_dop / (resa / 100)
    return totale_latte


def _cagliata_conferitori(client, caseificio_id, data_giorno):
    conf_b = _conferitori_attivi_categoria(client, caseificio_id, ["allevatore", "caseificio", "intermediario"], "semilavorato_bufala")
    conf_v = _conferitori_attivi_categoria(client, caseificio_id, ["allevatore", "caseificio", "intermediario"], "semilavorato_vaccino")
    return _righe_dati_categoria(client, conf_b, data_giorno), _righe_dati_categoria(client, conf_v, data_giorno)


def _ricotta_e_mista_e_congelamento_uscita(client, caseificio_id, data_giorno):
    kg_buf_mista = registro_calc.mista_consumato(client, caseificio_id, "bufala", data_giorno)
    kg_vac_mista = registro_calc.mista_consumato(client, caseificio_id, "vaccino", data_giorno)
    movimenti_uscita = (
        client.table("movimenti_congelato").select("*")
        .eq("caseificio_id", caseificio_id).eq("data", str(data_giorno)).eq("tipo", "congelamento")
        .execute().data
    )
    kg_congelamento_uscita = sum(float(m.get("kg") or 0) for m in movimenti_uscita)
    ddt_congelamento_uscita = ", ".join(m["ddt"] for m in movimenti_uscita if m.get("ddt"))
    return kg_buf_mista, kg_vac_mista, kg_congelamento_uscita, ddt_congelamento_uscita


def _vendite_giorno(client, caseificio_id, data_giorno):
    try:
        return (
            client.table("vendite_latte_destinatari")
            .select("*, destinatari_vendita(ragione_sociale)")
            .eq("caseificio_id", caseificio_id).eq("data", str(data_giorno))
            .execute().data
        )
    except Exception:
        return []


# ------------------------------------------------------------
# BLOCCO: SCRITTURA (nuovo - nessun insert_rows/delete_rows, nessuna fusione
# a rischio: si scrive sempre e solo in AVANTI, mai modificando righe gia'
# scritte, quindi le fusioni create qui non vengono mai toccate da un
# inserimento successivo)
# ------------------------------------------------------------
def _titolo(ws, riga, testo, sottotesto):
    ws.merge_cells(start_row=riga, start_column=1, end_row=riga, end_column=N_COLONNE)
    c = ws.cell(row=riga, column=1, value=testo)
    c.font = FONT_TITOLO
    riga += 1
    ws.merge_cells(start_row=riga, start_column=1, end_row=riga, end_column=N_COLONNE)
    c = ws.cell(row=riga, column=1, value=sottotesto)
    c.font = FONT_SOTTOTITOLO_INFO
    return riga + 2


def _sezione(ws, riga, testo):
    ws.merge_cells(start_row=riga, start_column=1, end_row=riga, end_column=N_COLONNE)
    c = ws.cell(row=riga, column=1, value=f"  {testo}")
    c.font = FONT_SEZIONE
    c.fill = FILL_SEZIONE
    for col in range(1, N_COLONNE + 1):
        ws.cell(row=riga, column=col).fill = FILL_SEZIONE
    return riga + 1


def _sottosezione(ws, riga, testo):
    c = ws.cell(row=riga, column=1, value=testo)
    c.font = FONT_SOTTOSEZIONE
    return riga + 1


def _tabella_conferitori(ws, riga, righe_dati):
    """Intestazione + righe (Provenienza | Cod. ASL/P.IVA | DDT | KG | Acidita' | Temperatura |
    Esito). Se righe_dati e' vuota, scrive solo una riga "(nessun conferitore attivo)"."""
    intestazioni = ["Provenienza", "Cod. ASL / P.IVA", "N. DDT", "KG", "Acidità", "Temperatura", "Esito controlli"]
    for col, testo in enumerate(intestazioni, start=1):
        c = ws.cell(row=riga, column=col, value=testo)
        c.font = FONT_INTESTAZIONE
        c.fill = FILL_INTESTAZIONE
        c.alignment = ALLINEA_CENTRO
        c.border = BORDO_INTESTAZIONE
    riga += 1
    if not righe_dati:
        c = ws.cell(row=riga, column=1, value="(nessun conferitore attivo in questa categoria)")
        c.font = FONT_VUOTO
        return riga + 1
    for d in righe_dati:
        ws.cell(row=riga, column=1, value=d["provenienza"])
        ws.cell(row=riga, column=2, value=d["codice"])
        ws.cell(row=riga, column=3, value=d["ddt"])
        ws.cell(row=riga, column=4, value=d["kg"])
        ws.cell(row=riga, column=5, value=_acidita_random() if d["kg"] > 0 else "")
        ws.cell(row=riga, column=6, value=_temperatura_random() if d["kg"] > 0 else "")
        ws.cell(row=riga, column=7, value="OK" if d["kg"] > 0 else "")
        riga += 1
    return riga


def _riga_valore(ws, riga, etichetta, valore, unita="kg"):
    ws.cell(row=riga, column=1, value=etichetta)
    ws.cell(row=riga, column=2, value=round(valore) if valore else 0)
    ws.cell(row=riga, column=3, value=unita)
    return riga + 1


def _riga_totale(ws, riga, etichetta, valore):
    for col in range(1, 4):
        ws.cell(row=riga, column=col).fill = FILL_TOTALE
    c1 = ws.cell(row=riga, column=1, value=etichetta)
    c1.font = FONT_TOTALE
    c2 = ws.cell(row=riga, column=2, value=round(valore) if valore else 0)
    c2.font = FONT_TOTALE
    ws.cell(row=riga, column=3, value="kg").font = FONT_TOTALE
    return riga + 1


def _compila_tr(ws, client, caseificio_id, data_giorno, prodotti_ammessi=None):
    if prodotti_ammessi is None:
        prodotti_ammessi = _prodotti_ammessi_nel_periodo(client, caseificio_id, data_giorno, data_giorno)

    for i, w in enumerate([26, 16, 14, 12, 14, 16, 14], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    anagrafica, is_dop = _caseificio_is_dop(client, caseificio_id)
    riga = _titolo(
        ws, 1,
        f"TR - Tabellone giornaliero rintracciabilità — {anagrafica['ragione_sociale']}",
        f"Data: {data_giorno.strftime('%d/%m/%Y')}" + ("  •  Caseificio DOP" if is_dop else ""),
    )

    totali = {"bufala_dop": 0.0, "bufala": 0.0, "vaccino": 0.0}

    # ------------------------------------------------------------
    # SEZIONE 1: LATTE IN INGRESSO
    # ------------------------------------------------------------
    riga = _sezione(ws, riga, "LATTE IN INGRESSO")

    if is_dop:
        riga = _sottosezione(ws, riga, "Allevamenti — Bufala DOP")
        conf = _conferitori_attivi_categoria(client, caseificio_id, ["allevatore"], "bufala_dop")
        righe_d = _righe_dati_categoria(client, conf, data_giorno)
        totali["bufala_dop"] += sum(r["kg"] for r in righe_d)
        riga = _tabella_conferitori(ws, riga, righe_d) + 1

        riga = _sottosezione(ws, riga, "Caseifici — Bufala DOP")
        conf = _conferitori_attivi_categoria(client, caseificio_id, ["caseificio"], "bufala_dop")
        righe_d = _righe_dati_categoria(client, conf, data_giorno)
        totali["bufala_dop"] += sum(r["kg"] for r in righe_d)
        riga = _tabella_conferitori(ws, riga, righe_d) + 1

    riga = _sottosezione(ws, riga, "Caseifici — Bufala non-DOP")
    conf = _conferitori_attivi_categoria(client, caseificio_id, ["caseificio"], "bufala")
    righe_d = _righe_dati_categoria(client, conf, data_giorno)
    totali["bufala"] += sum(r["kg"] for r in righe_d)
    riga = _tabella_conferitori(ws, riga, righe_d) + 1

    riga = _sottosezione(ws, riga, "Allevamenti — Bufala non-DOP")
    conf = _conferitori_attivi_categoria(client, caseificio_id, ["allevatore"], "bufala")
    righe_d = _righe_dati_categoria(client, conf, data_giorno)
    totali["bufala"] += sum(r["kg"] for r in righe_d)
    riga = _tabella_conferitori(ws, riga, righe_d) + 1

    riga = _sottosezione(ws, riga, "Latte vaccino")
    conf = _conferitori_attivi_categoria(client, caseificio_id, ["allevatore", "caseificio", "intermediario"], "vaccino")
    righe_d = _righe_dati_categoria(client, conf, data_giorno)
    totali["vaccino"] += sum(r["kg"] for r in righe_d)
    riga = _tabella_conferitori(ws, riga, righe_d) + 1

    riga = _sottosezione(ws, riga, "Congelato — latte scongelato rientrato in produzione (bufala non-DOP)")
    righe_congelato = _righe_congelato(client, caseificio_id, data_giorno)
    totali["bufala"] += sum(r["kg"] for r in righe_congelato)
    if righe_congelato:
        riga = _tabella_conferitori(ws, riga, righe_congelato) + 1
    else:
        c = ws.cell(row=riga, column=1, value="(nessuno scongelamento avvenuto oggi)")
        c.font = FONT_VUOTO
        riga += 2

    conf_cag_b, conf_cag_v = _cagliata_conferitori(client, caseificio_id, data_giorno)
    kg_cagliata_b = sum(r["kg"] for r in conf_cag_b)
    kg_cagliata_v = sum(r["kg"] for r in conf_cag_v)
    riga = _sottosezione(ws, riga, "Cagliata bufala ricevuta (semilavorato — esclusa dai totali latte)")
    riga = _tabella_conferitori(ws, riga, conf_cag_b) + 1
    riga = _sottosezione(ws, riga, "Cagliata vaccina ricevuta (semilavorato — esclusa dai totali latte)")
    riga = _tabella_conferitori(ws, riga, conf_cag_v) + 1

    if is_dop:
        riga = _riga_totale(ws, riga, "TOTALE latte bufala DOP in ingresso", totali["bufala_dop"])
    riga = _riga_totale(ws, riga, "TOTALE latte bufala non-DOP in ingresso (incl. congelato rientrato)", totali["bufala"])
    riga = _riga_totale(ws, riga, "TOTALE latte vaccino in ingresso", totali["vaccino"])
    riga += 1

    if is_dop:
        giacenza_apertura_dop = registro_calc.giacenza_apertura(client, caseificio_id, "bufala_dop", data_giorno)
        riga = _riga_valore(ws, riga, "Giacenza di apertura latte bufala DOP (dal Registro)", giacenza_apertura_dop)
        riga += 1

    # ------------------------------------------------------------
    # SEZIONE 2: LAVORAZIONE
    # ------------------------------------------------------------
    riga = _sezione(ws, riga, "LAVORAZIONE — LATTE TRASFORMATO")

    resa = _resa_dop_giorno(client, caseificio_id, data_giorno) if is_dop else None
    if is_dop:
        kg_buf_dop = registro_calc.trasformato(client, caseificio_id, "bufala_dop", data_giorno)
        riga = _riga_valore(ws, riga, "Latte bufala DOP trasformato", kg_buf_dop)
        kg_declassato = _latte_dop_per_gruppo(client, caseificio_id, data_giorno, resa, is_dop_prodotto=False)
        riga = _riga_valore(ws, riga, "...di cui latte DOP usato per prodotto declassato (non-DOP)", kg_declassato)
        kg_delattosata = _latte_dop_per_gruppo(client, caseificio_id, data_giorno, resa, is_dop_prodotto=True, solo_nome="senza lattosio")
        riga = _riga_valore(ws, riga, "...di cui latte DOP usato per mozzarella delattosata", kg_delattosata)

    kg_buf_non_dop = registro_calc.trasformato(client, caseificio_id, "bufala", data_giorno)
    riga = _riga_valore(ws, riga, "Latte bufala non-DOP trasformato", kg_buf_non_dop)
    kg_buf_mista, kg_vac_mista, kg_cong_uscita, ddt_cong_uscita = _ricotta_e_mista_e_congelamento_uscita(client, caseificio_id, data_giorno)
    riga = _riga_valore(ws, riga, "...di cui latte bufala usato per la Mozzarella Mista", kg_buf_mista)
    kg_vaccino_trasf = registro_calc.trasformato(client, caseificio_id, "vaccino", data_giorno)
    riga = _riga_valore(ws, riga, "Latte vaccino trasformato", kg_vaccino_trasf)
    riga = _riga_valore(ws, riga, "...di cui latte vaccino usato per la Mozzarella Mista", kg_vac_mista)
    riga = _riga_valore(ws, riga, "Cagliata bufala prodotta oggi", kg_cagliata_b)
    riga = _riga_valore(ws, riga, "Cagliata vaccina prodotta oggi", kg_cagliata_v)
    riga += 1

    riga = _sottosezione(ws, riga, "Latte destinato al congelamento (uscita)")
    ws.cell(row=riga, column=1, value="KG inviati a congelare")
    ws.cell(row=riga, column=2, value=round(kg_cong_uscita) if kg_cong_uscita else 0)
    ws.cell(row=riga, column=3, value=f"DDT: {ddt_cong_uscita}" if ddt_cong_uscita else "")
    riga += 2

    # ------------------------------------------------------------
    # SEZIONE 3: LATTE VENDUTO (nessun tetto massimo)
    # ------------------------------------------------------------
    riga = _sezione(ws, riga, "LATTE VENDUTO")
    vendite = _vendite_giorno(client, caseificio_id, data_giorno)
    intestazioni_vendite = ["Destinatario", "Tipo latte", "KG", "N. DDT", "", "", ""]
    for col, testo in enumerate(intestazioni_vendite[:4], start=1):
        c = ws.cell(row=riga, column=col, value=testo)
        c.font = FONT_INTESTAZIONE
        c.fill = FILL_INTESTAZIONE
        c.border = BORDO_INTESTAZIONE
    riga += 1
    if not vendite:
        c = ws.cell(row=riga, column=1, value="(nessuna vendita di latte oggi)")
        c.font = FONT_VUOTO
        riga += 1
    else:
        for v in vendite:
            ws.cell(row=riga, column=1, value=(v.get("destinatari_vendita") or {}).get("ragione_sociale", ""))
            ws.cell(row=riga, column=2, value=v.get("tipo_latte", ""))
            ws.cell(row=riga, column=3, value=round(float(v.get("kg") or 0)))
            ws.cell(row=riga, column=4, value=v.get("ddt") or "")
            riga += 1
    riga += 1

    # ------------------------------------------------------------
    # SEZIONE 4: GIACENZA DI CHIUSURA
    # ------------------------------------------------------------
    riga = _sezione(ws, riga, "GIACENZA DI CHIUSURA (dal Registro)")
    if is_dop:
        riga = _riga_valore(ws, riga, "Latte bufala DOP", registro_calc.giacenza_chiusura(client, caseificio_id, "bufala_dop", data_giorno))
    riga = _riga_valore(ws, riga, "Latte bufala non-DOP", registro_calc.giacenza_chiusura(client, caseificio_id, "bufala", data_giorno))
    riga += 1

    # ------------------------------------------------------------
    # SEZIONE 5: PRODOTTI OTTENUTI (solo quelli fatti nel periodo selezionato)
    # ------------------------------------------------------------
    riga = _sezione(ws, riga, "PRODOTTI OTTENUTI")
    prodotti = _prodotti_finiti(client, prodotti_ammessi, data_giorno)
    intestazioni_prodotti = ["Prodotto", "KG totale", "Vendita diretta", "Vendita a terzi", "", "", ""]
    for col, testo in enumerate(intestazioni_prodotti[:4], start=1):
        c = ws.cell(row=riga, column=col, value=testo)
        c.font = FONT_INTESTAZIONE
        c.fill = FILL_INTESTAZIONE
        c.border = BORDO_INTESTAZIONE
    riga += 1
    if not prodotti:
        c = ws.cell(row=riga, column=1, value="(nessun prodotto fatto nel periodo selezionato)")
        c.font = FONT_VUOTO
        riga += 1
    else:
        for p in prodotti:
            ws.cell(row=riga, column=1, value=p["nome"])
            ws.cell(row=riga, column=2, value=p["totale"])
            ws.cell(row=riga, column=3, value=p["diretta"])
            ws.cell(row=riga, column=4, value=p["terzi"])
            riga += 1


def genera_tr(client, caseificio_id, data_giorno, output_path):
    """Un solo giorno = un solo file, un solo foglio 'tr' costruito da zero."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = FOGLIO
    _compila_tr(ws, client, caseificio_id, data_giorno)
    wb.save(output_path)
    return output_path


def genera_tr_periodo(client, caseificio_id, data_da, data_a, output_path):
    """Un file con UN FOGLIO PER GIORNO nell'intervallo [data_da, data_a] (inclusi), ciascuno
    costruito da zero (nessun rischio di corruzione tra un giorno e l'altro, dato che ogni
    foglio e' indipendente e non copia mai un foglio gia' scritto)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    prodotti_ammessi = _prodotti_ammessi_nel_periodo(client, caseificio_id, data_da, data_a)

    n_giorni = (data_a - data_da).days + 1
    for i in range(n_giorni):
        giorno = data_da + _dt.timedelta(days=i)
        ws = wb.create_sheet(title=giorno.strftime("%d-%m-%Y"))
        _compila_tr(ws, client, caseificio_id, giorno, prodotti_ammessi=prodotti_ammessi)

    wb.save(output_path)
    return output_path
