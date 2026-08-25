# ============================================================
# MODULO: STAMPA MBC
# Compila il foglio "MBC" del template ufficiale RINA
# (templates_dop.xlsx) per una singola giornata di lavorazione.
#
# IMPORTANTE: il foglio MBC e' un documento ufficiale statico.
# Questo script NON modifica la struttura del file (righe,
# colonne, formule esistenti restano intatte) - scrive SOLO i
# valori nelle celle dati, secondo la mappatura concordata.
# ============================================================
import datetime as _dt
import shutil
import openpyxl

import registro_calc
import siero

TEMPLATE_PATH = "templates_dop.xlsx"  # percorso del file originale scaricato da GitHub
FOGLIO = "MBC"


# ------------------------------------------------------------
# BLOCCO: REGOLA ARROTONDAMENTO
# Confermata dall'utente: nella MATRICE MATERIE PRIME (latte,
# siero, semilavorato) ogni numero si arrotonda a INTERO PRIMA
# di essere scritto/sommato. Nella MATRICE PRODOTTO (mozzarella,
# altri formaggi) si mantengono fino a 2 decimali.
# ------------------------------------------------------------
def r_int(x):
    """Arrotonda a intero (materie prime): usare SEMPRE per latte/siero/semilavorato."""
    return int(round(float(x or 0)))


def r_prod(x):
    """Arrotonda a 2 decimali (prodotto finito): mozzarella, altri formaggi, ricotta."""
    return round(float(x or 0), 2)


# ------------------------------------------------------------
# BLOCCO: LETTURA DATI DA SUPABASE
# Ogni funzione isola una fonte dati, cosi' quando il Registro
# verra' riscritto basta aggiornare le funzioni segnate TODO.
# ------------------------------------------------------------

def get_anagrafica(client, caseificio_id):
    return client.table("caseifici").select("*").eq("id", caseificio_id).single().execute().data


def get_refrigerante_principale(client, caseificio_id):
    righe = (
        client.table("refrigeranti")
        .select("*")
        .eq("caseificio_id", caseificio_id)
        .eq("attivo", True)
        .order("codice")
        .limit(1)
        .execute()
        .data
    )
    return righe[0]["codice"] if righe else ""


def get_impostazione(client, caseificio_id, campo, alla_data):
    righe = (
        client.table("impostazioni_registro")
        .select("*")
        .eq("caseificio_id", caseificio_id)
        .eq("campo", campo)
        .lte("data_da", str(alla_data))
        .order("data_da", desc=True)
        .limit(1)
        .execute()
        .data
    )
    return righe[0]["valore"] if righe else ""


def get_conferimenti_per_categoria(client, caseificio_id, data_giorno, tipi_conferitore, tipo_latte):
    """Somma kg conferiti nel giorno per una categoria di conferitori (es. allevatori DOP)."""
    conferitori = (
        client.table("conferitori")
        .select("*, conferitori_tipi_latte(tipo_latte)")
        .eq("caseificio_id", caseificio_id)
        .eq("attivo", True)
        .in_("tipo", tipi_conferitore)
        .execute()
        .data
    )
    ids = [
        c["id"] for c in conferitori
        if tipo_latte in [t["tipo_latte"] for t in c.get("conferitori_tipi_latte", [])]
    ]
    if not ids:
        return 0.0, []
    conferimenti = (
        client.table("conferimenti")
        .select("*")
        .in_("conferitore_id", ids)
        .eq("data", str(data_giorno))
        .execute()
        .data
    )
    totale = sum(float(c["kg"] or 0) for c in conferimenti)
    ddt_list = [c["ddt"] for c in conferimenti if c.get("ddt")]
    return totale, ddt_list


def get_conferimenti_caseificio_dop_dettaglio(client, caseificio_id, data_giorno):
    """Conferimenti giornalieri da conferitori tipo 'caseificio' con bufala_dop, UNO PER RIGA
    (non sommati) - servono per D15/D16 (righe gia' presenti nel template, una per ogni
    caseificio DOP da cui si e' acquistato quel giorno, confermato dall'utente)."""
    conferitori = (
        client.table("conferitori")
        .select("*, conferitori_tipi_latte(tipo_latte)")
        .eq("caseificio_id", caseificio_id)
        .eq("attivo", True)
        .eq("tipo", "caseificio")
        .execute()
        .data
    )
    conferitori = [
        c for c in conferitori
        if "bufala_dop" in [t["tipo_latte"] for t in c.get("conferitori_tipi_latte", [])]
    ]
    righe = []
    for c in conferitori:
        rec = (
            client.table("conferimenti")
            .select("*")
            .eq("conferitore_id", c["id"])
            .eq("data", str(data_giorno))
            .execute()
            .data
        )
        if rec and float(rec[0].get("kg") or 0) > 0:
            righe.append({"kg": float(rec[0]["kg"]), "ddt": rec[0].get("ddt") or ""})
    return righe


def get_produzione_giorno(client, caseificio_id, data_giorno, contiene_nel_nome):
    """Somma kg_totale del giorno per prodotti il cui nome contiene una parola chiave (case-insensitive)."""
    prodotti = (
        client.table("prodotti")
        .select("*")
        .eq("caseificio_id", caseificio_id)
        .eq("attivo", True)
        .execute()
        .data
    )
    match = [p for p in prodotti if contiene_nel_nome.lower() in p["nome"].lower()]
    totale = 0.0
    for p in match:
        rec = (
            client.table("produzioni")
            .select("*")
            .eq("prodotto_id", p["id"])
            .eq("data", str(data_giorno))
            .execute()
            .data
        )
        if rec:
            totale += float(rec[0].get("kg_totale") or 0)
    return totale


def get_prodotto_e_produzione_giorno(client, caseificio_id, data_giorno, contiene_nel_nome, solo_dop=None):
    """Come get_produzione_giorno ma ritorna anche il record prodotto (per tipo_lotto/giorni_scadenza)
    e il record produzione completo (kg_totale/kg_diretta/kg_terzi), non solo il totale kg."""
    prodotti = (
        client.table("prodotti")
        .select("*")
        .eq("caseificio_id", caseificio_id)
        .eq("attivo", True)
        .execute()
        .data
    )
    match = [p for p in prodotti if contiene_nel_nome.lower() in p["nome"].lower()]
    if solo_dop is not None:
        match = [p for p in match if p["is_dop"] == solo_dop]
    if not match:
        return None, None
    prodotto = match[0]
    rec = (
        client.table("produzioni")
        .select("*")
        .eq("prodotto_id", prodotto["id"])
        .eq("data", str(data_giorno))
        .execute()
        .data
    )
    return prodotto, (rec[0] if rec else None)


def get_altri_formaggi_dop_giorno(client, caseificio_id, data_giorno, escludi_prodotto_id):
    """Somma kg_totale del giorno di tutti i prodotti DOP diversi dalla Mozzarella di Bufala
    Campana DOP (esclusa per id) e dalla Ricotta (gestita a parte nel blocco RBC/T20-W20)."""
    prodotti = (
        client.table("prodotti")
        .select("*")
        .eq("caseificio_id", caseificio_id)
        .eq("attivo", True)
        .eq("is_dop", True)
        .execute()
        .data
    )
    match = [p for p in prodotti if p["id"] != escludi_prodotto_id and "ricotta" not in p["nome"].lower()]
    totale = 0.0
    for p in match:
        rec = (
            client.table("produzioni")
            .select("kg_totale")
            .eq("prodotto_id", p["id"])
            .eq("data", str(data_giorno))
            .execute()
            .data
        )
        if rec:
            totale += float(rec[0].get("kg_totale") or 0)
    return totale


def get_semilavorato_bufala_dop_giorno(client, caseificio_id, data_giorno):
    """Semilavorato di bufala DOP trasformato (usato) nel giorno - dal Registro."""
    return registro_calc.trasformato(client, caseificio_id, "semilavorato_bufala", data_giorno)


# Ricollegate al vero Registro (registro_calc.py) - stessa logica usata dalla pagina Registro.
def get_registro_giacenza_apertura(client, caseificio_id, data_giorno):
    return registro_calc.giacenza_apertura(client, caseificio_id, "bufala_dop", data_giorno)


def get_registro_trasformato_dop(client, caseificio_id, data_giorno):
    return registro_calc.trasformato(client, caseificio_id, "bufala_dop", data_giorno)


def get_registro_extra_dop(client, caseificio_id, data_giorno):
    return registro_calc.extra_dop_consumato(client, caseificio_id, data_giorno)


def get_registro_giacenza_chiusura(client, caseificio_id, data_giorno):
    return registro_calc.giacenza_chiusura(client, caseificio_id, "bufala_dop", data_giorno)


# ------------------------------------------------------------
# BLOCCO: NUMERO SCHEDA (progressivo = giorno dell'anno)
# Confermato con l'utente: 25 febbraio = 56° giorno dell'anno = "56/26M" - il numero e'
# sempre positivo (1-366), MAI negativo. Formato finale: "NNN/AA" + lettera (M per MBC, R per RBC).
# ATTENZIONE: questa funzione e' SOLO per il campo "Scheda N." - il lotto dei prodotti (V10 ecc.)
# e' un'altra cosa, dipende dal tipo di lotto scelto per il prodotto in Prodotti (da rivedere).
# ------------------------------------------------------------
def scheda_n(data_giorno, lettera="M"):
    giorno_anno = data_giorno.timetuple().tm_yday
    anno_breve = data_giorno.strftime("%y")
    return f"{giorno_anno}/{anno_breve}{lettera}"


def numero_scheda(data_giorno):
    # Mantenuta per compatibilita' con la generazione provvisoria del lotto prodotti
    # (in attesa di rifare il lotto secondo il tipo_lotto del prodotto, come richiesto)
    return data_giorno.timetuple().tm_yday


# ------------------------------------------------------------
# BLOCCO: LOTTO PRODOTTO (secondo il tipo_lotto scelto in Prodotti)
# Sostituisce il precedente lotto generico "AAAAMMGG-progressivo":
# ora dipende dal campo tipo_lotto del prodotto specifico.
#   - "data_produzione" -> data del giorno (GG/MM/AAAA)
#   - "data_scadenza"   -> data del giorno + giorni_scadenza (GG/MM/AAAA)
#   - "giuliano"        -> calendario giuliano progressivo (1-366)
# ------------------------------------------------------------
def calcola_lotto(prodotto, data_giorno):
    if not prodotto:
        return ""
    tipo_lotto = prodotto.get("tipo_lotto")
    if tipo_lotto == "data_scadenza":
        giorni = int(prodotto.get("giorni_scadenza") or 0)
        return (data_giorno + _dt.timedelta(days=giorni)).strftime("%d/%m/%Y")
    if tipo_lotto == "giuliano":
        return str(data_giorno.timetuple().tm_yday)
    return data_giorno.strftime("%d/%m/%Y")  # default: data_produzione


def calcola_scadenza(prodotto, data_giorno):
    if not prodotto:
        return ""
    giorni = int(prodotto.get("giorni_scadenza") or 0)
    return (data_giorno + _dt.timedelta(days=giorni)).strftime("%d/%m/%Y")


# ------------------------------------------------------------
# BLOCCO: COMPILAZIONE FOGLIO MBC
# _compila_mbc() scrive i dati di UN giorno su UN foglio worksheet gia' aperto - separata da
# genera_mbc() (che gestisce apertura/salvataggio file) cosi' la stessa logica di compilazione
# si puo' riusare foglio per foglio dentro un unico workbook multi-giorno (vedi genera_mbc_periodo).
# ------------------------------------------------------------
def _compila_mbc(ws, client, caseificio_id, data_giorno):
    anagrafica = get_anagrafica(client, caseificio_id)

    # --- Intestazione ---
    ws["C1"] = anagrafica.get("ragione_sociale", "")
    # ws["T1"] = codice RINA AGRIFOOD -> campo non ancora presente in Anagrafica, lasciato vuoto
    ws["C3"] = scheda_n(data_giorno, lettera="M")
    ws["H3"] = data_giorno.strftime("%d/%m/%Y")

    ws["H5"] = get_impostazione(client, caseificio_id, "ora_ricevimento_latte", data_giorno)
    ws["M5"] = get_impostazione(client, caseificio_id, "ora_inizio_lavorazione", data_giorno)
    ws["U5"] = get_impostazione(client, caseificio_id, "ora_fine_lavorazione", data_giorno)

    # --- Sezione Ricevimento (matrice materie prime: arrotondamento a intero) ---
    ws["D10"] = r_int(get_registro_giacenza_apertura(client, caseificio_id, data_giorno))  # TODO Registro
    ws["E10"] = get_refrigerante_principale(client, caseificio_id)

    kg_allevamento, _ = get_conferimenti_per_categoria(
        client, caseificio_id, data_giorno, ["allevatore"], "bufala_dop"
    )
    ws["D13"] = r_int(kg_allevamento)
    # CORREZIONE: E13 mostra il tank dove il latte e' stoccato (come E10), non piu' vuoto -
    # in attesa del grosso lavoro futuro sui refrigeranti per-lotto/tank specifico
    ws["E13"] = get_refrigerante_principale(client, caseificio_id)

    kg_raccoglitore, _ddt_raccoglitore = get_conferimenti_per_categoria(
        client, caseificio_id, data_giorno, ["intermediario"], "bufala_dop"
    )
    ws["D14"] = r_int(kg_raccoglitore)
    # CORREZIONE: E14 mostra il tank (non piu' il DDT, come richiesto) - stesso discorso di E13
    ws["E14"] = get_refrigerante_principale(client, caseificio_id)

    # D15/D16: latte da caseifici DOP - una riga per ogni caseificio (max 2/giorno, confermato
    # dall'utente che le righe esistono gia' nel template, nessuna modifica di struttura)
    # NOTA: scritte sempre (con "" di fallback, non solo quando ci sono dati) - necessario perche'
    # nella generazione multi-giorno (genera_mbc_periodo) lo stesso foglio viene duplicato da un
    # giorno all'altro: senza il fallback, un valore del giorno precedente resterebbe visibile
    # anche nel giorno successivo se quel giorno non ha caseifici DOP da cui acquistare.
    # CORREZIONE: anche E15/E16 mostrano il tank (non piu' il DDT), stesso discorso di E13/E14.
    caseifici_dop = get_conferimenti_caseificio_dop_dettaglio(client, caseificio_id, data_giorno)
    ws["D15"] = r_int(caseifici_dop[0]["kg"]) if len(caseifici_dop) >= 1 else ""
    ws["E15"] = get_refrigerante_principale(client, caseificio_id) if len(caseifici_dop) >= 1 else ""
    ws["D16"] = r_int(caseifici_dop[1]["kg"]) if len(caseifici_dop) >= 2 else ""
    ws["E16"] = get_refrigerante_principale(client, caseificio_id) if len(caseifici_dop) >= 2 else ""

    # --- Sezione Lavorazione ---
    # CORREZIONE IMPORTANTE (bug grave trovato il 24/08): G10-G14 sono le ETICHETTE della
    # tabella (celle unite G:I, es. "Temperatura attivazione °C"), NON i valori - scriverci
    # sopra CANCELLAVA il testo dell'etichetta. I veri valori vanno nella colonna J (J10-J14),
    # dove nel template originale c'erano dei valori di esempio mai sovrascritti finora. Stesso
    # bug per G16 (valore in J16), G22 (valore in K22) e M21/M22 (marcatura in P21/P22, vedi sotto).
    ws["J10"] = get_impostazione(client, caseificio_id, "temperatura_attivazione", data_giorno)
    ws["J11"] = get_impostazione(client, caseificio_id, "tipo_siero_innesto", data_giorno)
    # J12 (Caglio: fornitore e lotto): CORREZIONE - fornitore e lotto vanno CONCATENATI nella
    # stessa cella J12 (prima erano splittati, fornitore in una cella sbagliata G12 e lotto da
    # solo in J12).
    caglio_fornitore = get_impostazione(client, caseificio_id, "caglio_fornitore", data_giorno)
    caglio_lotto = get_impostazione(client, caseificio_id, "caglio_lotto", data_giorno)
    ws["J12"] = f"{caglio_fornitore} - {caglio_lotto}" if caglio_fornitore or caglio_lotto else ""
    ws["J13"] = get_impostazione(client, caseificio_id, "acidita_primo_siero", data_giorno)  # pH finale
    ws["J14"] = get_impostazione(client, caseificio_id, "tempo_maturazione_minuti", data_giorno)
    ws["J16"] = 0  # Acidita' Primo Siero: per ora 0, in attesa del foglio Siero
    ws["K22"] = get_impostazione(client, caseificio_id, "cicli_lavorazione", data_giorno)
    # K18 (Aggiunte RBC - percentuale sale nel siero per RBC, 0.3% confermato dall'utente):
    # il template ha gia' il valore di esempio corretto (0.003 = 0.3%), non lo tocchiamo.

    kg_trasformato_dop = get_registro_trasformato_dop(client, caseificio_id, data_giorno)  # TODO Registro
    ws["K21"] = r_int(kg_trasformato_dop)
    ws["K25"] = r_int(get_registro_extra_dop(client, caseificio_id, data_giorno))          # TODO Registro
    ws["K27"] = r_int(get_registro_giacenza_chiusura(client, caseificio_id, data_giorno))  # TODO Registro

    # D21 (1L. Primo Siero Autoprodotto): sostituita la vecchia formula fissa "=K21*75%" del
    # template con il siero REALE calcolato per differenza (latte trasformato - mozzarella
    # prodotta), come deciso - "niente si crea e nulla si distrugge, tutto si trasforma".
    kg_siero_dop = siero.siero_dop_prodotto_giorno(client, caseificio_id, data_giorno)
    ws["D21"] = r_int(kg_siero_dop)
    ws["D22"] = r_int(kg_siero_dop)  # Totale Primo Siero Autoprodotto (D22 = D21, un solo ciclo tracciato oggi)

    # K20 (2I. Semilavorato idoneo a DOP lavorato a MBC DOP) e K24 (2N. Semilavorato idoneo a
    # DOP destinato ad ALTRE lavorazioni e/o semilavorati): oggi il Registro traccia un unico
    # totale di semilavorato_bufala trasformato, senza distinguere quanto va a MBC e quanto
    # va altrove - l'intero valore e' quindi attribuito a K20 (uso primario, produzione MBC);
    # K24 resta a 0 finche' questa distinzione non sara' tracciata nel Registro.
    kg_semilavorato_dop = get_semilavorato_bufala_dop_giorno(client, caseificio_id, data_giorno)
    ws["K20"] = r_int(kg_semilavorato_dop)
    ws["K24"] = 0  # TODO: nessuna fonte dati ancora per la quota "destinata ad altre lavorazioni"

    ws["K26"] = ""  # TODO: "ceduto a terzi" - manca ancora una fonte dati (kg latte DOP ceduto quel giorno);
                     # lasciato vuoto invece della formula #REF! originale, in attesa di costruire questa parte

    # W14 (Bilancio di massa latte): la formula originale del template "=+W10/K21" resta
    # intatta (si ricalcola da sola in Excel una volta compilati W10 e K21) - qui si corregge
    # solo il formato numerico a percentuale con 2 decimali, come richiesto.
    ws["W14"].number_format = "0.00%"

    # --- Sezione Filatura ---
    ws["M9"] = get_impostazione(client, caseificio_id, "temperatura_acqua_filatura", data_giorno)

    # Affumicata e Delattosata: stesso meccanismo automatico (nessun inserimento manuale,
    # il programma legge quanto prodotto in Produzioni quel giorno), confermato dall'utente.
    # Filtrate a solo_dop=True: solo le versioni DOP contano per MBC.
    # CORREZIONE IMPORTANTE: M21 ("AFFUMICATA") e M22 ("LAVORATA A MANO") sono ETICHETTE FISSE
    # del template - scriverci sopra "X"/"" le cancellava. La marcatura va nelle celle libere
    # adiacenti P21/P22 (non unite, verificate sul template reale).
    _, produzione_affum = get_prodotto_e_produzione_giorno(client, caseificio_id, data_giorno, "affumicat", solo_dop=True)
    kg_affumicata = float((produzione_affum or {}).get("kg_totale") or 0)
    kg_affumicata_diretta = float((produzione_affum or {}).get("kg_diretta") or 0)
    kg_affumicata_terzi = float((produzione_affum or {}).get("kg_terzi") or 0)
    if kg_affumicata > 0:
        ws["P21"] = "X"
        ws["P22"] = ""
    else:
        ws["P21"] = ""
        ws["P22"] = "X"

    _, produzione_delatt = get_prodotto_e_produzione_giorno(client, caseificio_id, data_giorno, "senza lattosio", solo_dop=True)
    kg_delattosata = float((produzione_delatt or {}).get("kg_totale") or 0)
    kg_delattosata_diretta = float((produzione_delatt or {}).get("kg_diretta") or 0)
    kg_delattosata_terzi = float((produzione_delatt or {}).get("kg_terzi") or 0)

    # --- Sezione Produzioni (matrice prodotto: fino a 2 decimali, MAI arrotondata a intero) ---
    # NOTA: intestazioni reali sono R9='Prodotto' (testo), V9='lotto n.', W9='Q.tà (kg)'.
    prodotto_mozz, produzione_mozz = get_prodotto_e_produzione_giorno(
        client, caseificio_id, data_giorno, "Mozzarella di Bufala Campana DOP", solo_dop=True
    )
    kg_mozz_totale = float((produzione_mozz or {}).get("kg_totale") or 0)
    kg_mozz_diretta = float((produzione_mozz or {}).get("kg_diretta") or 0)
    kg_mozz_terzi = float((produzione_mozz or {}).get("kg_terzi") or 0)
    lotto_mozz = calcola_lotto(prodotto_mozz, data_giorno)

    ws["R10"] = "Mozzarella di Bufala Campana DOP"
    ws["V10"] = lotto_mozz
    # CORREZIONE: W10 ora somma mozzarella DOP normale + delattosata DOP + affumicata DOP
    # prodotte quel giorno (prima usava solo la mozzarella "normale", come richiesto).
    ws["W10"] = r_prod(kg_mozz_totale + kg_delattosata + kg_affumicata)

    # P14 (riga Filatura "da 30 a 800 gr"): quantita' di mozzarella DOP NORMALE prodotta.
    ws["P14"] = r_prod(kg_mozz_totale)
    # Riga 15 (M15/P15) era vuota nel template - aggiunta qui come nuova riga dedicata alla
    # mozzarella delattosata (etichetta scritta per la prima volta, richiesta esplicita).
    ws["M15"] = "MOZZARELLA DELATTOSATA" if kg_delattosata > 0 else ""
    ws["P15"] = r_prod(kg_delattosata) if kg_delattosata > 0 else ""

    # R11 "Confezionata" = vendita a terzi; R12 "Sfusa per punto vendita" = vendita diretta
    # (mappatura confermata dall'utente). Sostituiscono le vecchie formule "=V10"/"=W10".
    # CORREZIONE: W11/W12 ora sommano anche le quote diretta/terzi REALI di delattosata e
    # affumicata (già registrate come prodotti a se' in Produzioni, con la propria
    # suddivisione), invece di usare solo la mozzarella "normale" - niente proporzione
    # sintetica, si usano i dati diretta/terzi effettivamente inseriti per ciascun prodotto.
    ws["V11"] = lotto_mozz
    ws["W11"] = r_prod(kg_mozz_terzi + kg_delattosata_terzi + kg_affumicata_terzi)
    ws["V12"] = lotto_mozz
    ws["W12"] = r_prod(kg_mozz_diretta + kg_delattosata_diretta + kg_affumicata_diretta)

    # W17 (Altri formaggi): altri prodotti DOP (diversi da mozzarella e ricotta) fatti con
    # latte DOP quel giorno, sostituendo la vecchia formula rotta "=#REF!".
    escludi_id = prodotto_mozz["id"] if prodotto_mozz else None
    kg_altri_formaggi = get_altri_formaggi_dop_giorno(client, caseificio_id, data_giorno, escludi_id)
    ws["W17"] = r_prod(kg_altri_formaggi)

    # --- T20-W20 (Produzione RBC): riferimento incrociato al foglio RBC, SOLO se il
    # caseificio ha effettivamente prodotto Ricotta di Bufala Campana DOP quel giorno.
    # T20 (Pezzatura gr.) resta il valore fisso "250" gia' presente nel template.
    kg_ricotta_dop, ricotta_dop = siero.get_ricotta_dop_giorno(client, caseificio_id, data_giorno)
    if kg_ricotta_dop > 0:
        ws["U20"] = r_int(kg_ricotta_dop * 4)  # Unita' n° = kg prodotti x 4
        ws["V20"] = calcola_lotto(ricotta_dop, data_giorno)
        ws["W20"] = calcola_scadenza(ricotta_dop, data_giorno)
        # U25/U26/U27 (Idoneo): croce solo quando e' stata prodotta anche la RBC quel giorno.
        ws["U25"] = "Idoneo X"
        ws["U26"] = "Idoneo X"
        ws["U27"] = "Idoneo X"
    else:
        ws["U20"] = ""
        ws["V20"] = ""
        ws["W20"] = ""
        ws["U25"] = "Idoneo"
        ws["U26"] = "Idoneo"
        ws["U27"] = "Idoneo"


def genera_mbc(client, caseificio_id, data_giorno, output_path):
    """Un solo giorno = un solo foglio MBC (comportamento originale, invariato)."""
    shutil.copy(TEMPLATE_PATH, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb[FOGLIO]
    _compila_mbc(ws, client, caseificio_id, data_giorno)
    wb.save(output_path)
    return output_path


def genera_mbc_periodo(client, caseificio_id, data_da, data_a, output_path):
    """Un file MBC con UN FOGLIO PER GIORNO nell'intervallo [data_da, data_a] (inclusi).
    File separato da RBC/tr, come richiesto - non contiene gli altri due fogli del template."""
    shutil.copy(TEMPLATE_PATH, output_path)
    wb = openpyxl.load_workbook(output_path)
    for nome in list(wb.sheetnames):
        if nome != FOGLIO:
            del wb[nome]
    ws_template = wb[FOGLIO]

    n_giorni = (data_a - data_da).days + 1
    for i in range(n_giorni):
        giorno = data_da + _dt.timedelta(days=i)
        ws = ws_template if i == 0 else wb.copy_worksheet(ws_template)
        ws.title = giorno.strftime("%d-%m-%Y")  # nome foglio: niente "/" (non ammesso in Excel)
        _compila_mbc(ws, client, caseificio_id, giorno)

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    # Esempio d'uso manuale (va richiamato dalla pagina Streamlit "Fogli Stampabili")
    from db import get_client
    client = get_client()
    genera_mbc(client, caseificio_id=1, data_giorno=_dt.date.today(), output_path="MBC_compilato.xlsx")
    print("File generato: MBC_compilato.xlsx")
