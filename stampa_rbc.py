# ============================================================
# MODULO: STAMPA RBC
# Compila il foglio "RBC" del template ufficiale RINA
# (templates_dop.xlsx) per una singola giornata di lavorazione.
#
# IMPORTANTE: il foglio RBC e' un documento ufficiale statico,
# come MBC - la struttura del file non viene mai modificata.
#
# NOTA: Scheda N. (Q5) e Data (U5) sul foglio RBC sono gia'
# collegate con formula al foglio MBC dello stesso workbook
# (=MBC!C3 e =MBC!H3) - non serve scriverle qui, basta che
# genera_mbc() sia stata eseguita sullo stesso file prima
# (o nello stesso momento) di genera_rbc().
# ============================================================
import datetime as _dt
import math
import shutil
import openpyxl

import siero
from stampa_mbc import (
    get_anagrafica, get_impostazione, get_produzione_giorno, numero_scheda, scheda_n,
    calcola_lotto, r_int,
)

FOGLIO = "RBC"

# Righe reali della tabella "Primo Siero Acquistato" (verificate sul template: la riga 11 e'
# l'intestazione A11/E11/I11/M11/S11, i dati veri vanno nelle 4 righe sottostanti).
RIGHE_PS_ACQUISTATO = [13, 15, 17, 19]


# ------------------------------------------------------------
# BLOCCO: SERIALIZZAZIONE VALORI "SPECIALI" (uguale a quella usata in
# pages/7_Impostazioni_Fisse.py per siero_stabilizzato e trattamento_termico_ricotta:
# "SI|Opzione1,Opzione2" oppure "NO"). Duplicata qui (non importabile da una pagina Streamlit).
# ------------------------------------------------------------
def _parse_speciale(valore):
    if not valore or valore == "NO":
        return False, []
    if valore.startswith("SI|"):
        return True, [s for s in valore[3:].split(",") if s]
    return False, []


def _scadenza_corta(prodotto, data_giorno):
    """Come calcola_scadenza ma con anno a 2 cifre (GG/MM/AA) - formato richiesto specificamente
    per RBC T78, diverso dalla regola generale GG/MM/AAAA usata nel resto del programma."""
    if not prodotto:
        return ""
    giorni = int(prodotto.get("giorni_scadenza") or 0)
    return (data_giorno + _dt.timedelta(days=giorni)).strftime("%d/%m/%y")


# ------------------------------------------------------------
# BLOCCO: DATI DAL FOGLIO SIERO (acquisto/autoprodotto multi-giorno non ancora costruiti nel
# gestionale). Placeholder isolati, da ricollegare quando la fase Siero sara' pronta.
# ------------------------------------------------------------
def get_siero_acquistato(client, caseificio_id, data_giorno):
    """Ritorna una lista di acquisti (max 4/giorno, una riga per caseificio DOP - righe 13-19
    del template). Per ora nessuna fonte dati -> lista vuota, nessuna riga scritta."""
    # TODO: collegare al foglio Siero quando pronto
    return []  # es. futuro: [{"origine": ..., "ddt": ..., "kg": ..., "ora_rottura": ..., "tank": ...}, ...]


# ------------------------------------------------------------
# BLOCCO: COMPILAZIONE FOGLIO RBC
# _compila_rbc() scrive i dati di UN giorno su UN foglio worksheet gia' aperto - stessa idea di
# _compila_mbc() in stampa_mbc.py, per riutilizzare la logica dentro un workbook multi-giorno.
# Puo' essere chiamata sia sullo STESSO file su cui e' gia' stata chiamata genera_mbc() per lo
# stesso giorno (comportamento storico), sia in un file separato (vedi genera_rbc_periodo) - in
# entrambi i casi Q5/U5 vengono scritti come valori letterali, mai piu' come formula verso MBC.
# ------------------------------------------------------------
def _compila_rbc(ws, client, caseificio_id, data_giorno):
    anagrafica = get_anagrafica(client, caseificio_id)

    # --- Intestazione ---
    ws["C5"] = anagrafica.get("ragione_sociale", "")
    # ws["L5"] = codice RINA AGRIFOOD -> non ancora presente in Anagrafica, lasciato vuoto
    # CORREZIONE: Q5 NON resta piu' la formula =MBC!C3 (copierebbe anche la lettera "M" di MBC) -
    # va sovrascritta con la stessa data/progressivo ma lettera "R" (formato confermato: "31/26R")
    ws["Q5"] = scheda_n(data_giorno, lettera="R")
    # CORREZIONE: U5 (Data) NON resta piu' la formula originale "=MBC!H3" - ora che RBC puo'
    # essere generato anche in un file separato da MBC (fogli stampabili per periodo), quella
    # formula punterebbe a un foglio inesistente. Scritta come valore letterale, sempre corretto
    # sia nel file combinato MBC+RBC sia nel file RBC-solo del periodo.
    ws["U5"] = data_giorno.strftime("%d/%m/%Y")

    # --- Primo Siero Acquistato (righe dati reali 13/15/17/19 - la riga 11 e' l'intestazione,
    # errore corretto qui: la versione precedente scriveva sopra le etichette in riga 11).
    # Tutte le 4 righe vengono sempre scritte (anche vuote "") - necessario per la generazione
    # multi-giorno (genera_rbc_periodo), dove lo stesso foglio e' duplicato da un giorno
    # all'altro: senza pulizia esplicita, i dati di un giorno resterebbero visibili anche nel
    # foglio del giorno successivo se quel giorno non ha acquisti.
    acquistati = get_siero_acquistato(client, caseificio_id, data_giorno)
    for i, riga in enumerate(RIGHE_PS_ACQUISTATO):
        if i < len(acquistati):
            a = acquistati[i]
            ws[f"A{riga}"] = a.get("origine", "")
            ws[f"E{riga}"] = a.get("ddt", "")
            ws[f"I{riga}"] = r_int(a.get("kg", 0))
            ws[f"M{riga}"] = a.get("ora_rottura", "")
            ws[f"S{riga}"] = a.get("tank", "")
        else:
            ws[f"A{riga}"] = ""
            ws[f"E{riga}"] = ""
            ws[f"I{riga}"] = ""
            ws[f"M{riga}"] = ""
            ws[f"S{riga}"] = ""

    # --- Primo Siero Autoprodotto (riga 27 = SEMPRE il giorno stesso, come confermato) ---
    # F27/J27 restano le formule originali del template (=Q5 e =U5, cioe' scheda/data di
    # QUESTO stesso foglio RBC) - corrette di default, non si toccano.
    kg_siero_oggi = siero.siero_dop_prodotto_giorno(client, caseificio_id, data_giorno)
    ws["L27"] = r_int(kg_siero_oggi)  # stesso valore di MBC D22 (siero autoprodotto oggi)
    ws["P27"] = get_impostazione(client, caseificio_id, "ora_siero_autoprodotto_rbc", data_giorno)
    ws["T27"] = "trasformazione autoprodotta"
    # Righe 29/31/33: da usare quando la ricotta DOP del giorno consuma anche siero di giorni
    # PRECEDENTI (giacenza) - in quel caso vanno indicate scheda e data del giorno originale di
    # produzione di quel siero e T29/T31/T33 andrebbe scritto "tank siero". Non ancora costruito:
    # serve tracciare DA QUALI giorni specifici viene prelevato l'avanzo (oggi la giacenza in
    # siero.py e' solo un totale cumulativo, non un elenco di lotti per data). TODO futuro.

    # --- PS Stabilizzato (righe 37/38/39: Pastorizzato/Termizzato/Refrigerato) ---
    valore_stabilizzato = get_impostazione(client, caseificio_id, "siero_stabilizzato", data_giorno)
    si_stabilizzato, tipi_stabilizzato = _parse_speciale(valore_stabilizzato)
    ws["H37"] = "X" if si_stabilizzato and "Pastorizzato" in tipi_stabilizzato else ""
    ws["H38"] = "X" if si_stabilizzato and "Termizzato" in tipi_stabilizzato else ""
    ws["H39"] = "X" if si_stabilizzato and "Refrigerato" in tipi_stabilizzato else ""

    # --- Quantita' totale PS lavorato + n° cicli ---
    # H41: sostituita la formula rotta del template "=H78/4*40" - il PS lavorato e' il siero
    # REALMENTE consumato per la ricotta DOP prodotta oggi (ricotta / % resa, principio
    # "niente si crea e nulla si distrugge" gia' usato in siero.py).
    kg_ps_lavorato, _errore_resa = siero.siero_utilizzato_ricotta_dop_giorno(client, caseificio_id, data_giorno)
    ws["H41"] = r_int(kg_ps_lavorato)
    # N41: CORREZIONE - non piu' input manuale con default 1, calcolato dal siero lavorato:
    # ogni ciclo lavora al MASSIMO 900 kg (es. 1200 kg = 2 cicli, 2100 kg = 3 cicli).
    ws["N41"] = max(1, math.ceil(kg_ps_lavorato / 900)) if kg_ps_lavorato > 0 else 1

    # K45 (acidita' primo siero) resta la formula originale del template - non toccarla.

    # --- Aggiunte (percentuali da Impostazioni Fisse, convertite in kg sul PS lavorato) ---
    perc_latte_bufala = get_impostazione(client, caseificio_id, "perc_latte_bufala_rbc", data_giorno)
    perc_panna_fresca = get_impostazione(client, caseificio_id, "perc_panna_fresca_rbc", data_giorno)
    try:
        ws["K47"] = r_int(float(perc_latte_bufala or 0) / 100 * kg_ps_lavorato)
    except ValueError:
        ws["K47"] = ""
    try:
        ws["K49"] = r_int(float(perc_panna_fresca or 0) / 100 * kg_ps_lavorato)
    except ValueError:
        ws["K49"] = ""
    # K51 (sale) resta la formula originale del template "=H41*0.3%" - non toccarla (decisione
    # gia' presa in precedenza); il campo Impostazioni Fisse kg_sale_rbc resta disponibile per
    # un futuro riallineamento se si decidera' di sostituire anche questa formula.

    # --- Agenti acidificanti (F53/F55/F57 = "X" su quello scelto, gli altri vuoti) ---
    agente = get_impostazione(client, caseificio_id, "agente_acidificante_rbc", data_giorno)
    ws["F53"] = "X" if agente == "Cizza di Mozzarella di Bufala Campana DOP" else ""
    ws["F55"] = "X" if agente == "Acido Lattico" else ""
    ws["F57"] = "X" if agente == "Acido Citrico" else ""

    # --- Temperature ricotta (F59/F61, da Impostazioni Fisse - nuovi campi dedicati) ---
    temp_finale = get_impostazione(client, caseificio_id, "temperatura_finale_ricotta", data_giorno)
    temp_raffreddamento = get_impostazione(client, caseificio_id, "temperatura_raffreddamento_ricotta", data_giorno)
    ws["F59"] = f"{temp_finale} °C" if temp_finale else ""
    ws["F61"] = f"{temp_raffreddamento} °C" if temp_raffreddamento else ""

    # --- Trattamento termico della ricotta (SI/NO in L63/O63, Lisciatura/Omogeneizzazione in S63/T63) ---
    valore_trattamento = get_impostazione(client, caseificio_id, "trattamento_termico_ricotta", data_giorno)
    si_trattamento, tipi_trattamento = _parse_speciale(valore_trattamento)
    ws["L63"] = "x" if si_trattamento else ""
    ws["O63"] = "" if si_trattamento else "x"
    ws["S63"] = "x" if si_trattamento and "Lisciatura" in tipi_trattamento else ""
    # T63 non ha una cella separata per la croce nel template (finisce in colonna W) - la croce
    # viene aggiunta dentro la stessa etichetta, come unica soluzione compatibile con la struttura.
    ws["T63"] = "OMOGENEIZZAZIONE X" if si_trattamento and "Omogeneizzazione" in tipi_trattamento else "OMOGENEIZZAZIONE"

    # --- Caratteristiche prodotto finito ---
    # CORREZIONE IMPORTANTE: F68/I68 e R68/U68 sono le ETICHETTE delle colonne
    # ("Idoneo"/"Non Idoneo") - gia' presenti nel template, NON vanno sovrascritte.
    # Il valore va scritto in F70 (caratteristiche fisiche) e R70 (organolettiche),
    # con una "x" (come nel template originale), non con la parola "Idoneo".
    ws["F70"] = "x"
    ws["R70"] = "x"

    # --- Confezionamento (Ricotta di Bufala DOP prodotta) ---
    # D78 (Pezzatura): valore FISSO "250" gia' presente nel template - NON e' la quantita' di
    # ricotta prodotta, non va toccato (correzione: la versione precedente lo sovrascriveva per errore).
    # H78 (Unita' n°): kg prodotti MOLTIPLICATO 4. L78 (ID Lotto): secondo il tipo_lotto reale
    # del prodotto (stessa logica di MBC V10). T78 (Scadenza): formato GG/MM/AA (2 cifre anno,
    # richiesta specifica per questa cella, diversa dalla regola generale GG/MM/AAAA).
    kg_ricotta, prodotto_ricotta = siero.get_ricotta_dop_giorno(client, caseificio_id, data_giorno)
    if kg_ricotta > 0:
        ws["H78"] = r_int(kg_ricotta * 4)
        ws["L78"] = calcola_lotto(prodotto_ricotta, data_giorno)
        ws["P78"] = data_giorno.strftime("%d/%m/%Y")  # data confezionamento
        ws["T78"] = _scadenza_corta(prodotto_ricotta, data_giorno)
    else:
        ws["H78"] = ""
        ws["L78"] = ""
        ws["P78"] = ""
        ws["T78"] = ""


def genera_rbc(client, caseificio_id, data_giorno, output_path):
    """Un solo giorno = un solo foglio RBC (comportamento originale, invariato) - va chiamata
    sullo STESSO file .xlsx su cui e' gia' stata chiamata genera_mbc() per lo stesso giorno."""
    wb = openpyxl.load_workbook(output_path)
    ws = wb[FOGLIO]
    _compila_rbc(ws, client, caseificio_id, data_giorno)
    wb.save(output_path)
    return output_path


def genera_rbc_periodo(client, caseificio_id, data_da, data_a, output_path):
    """Un file RBC con UN FOGLIO PER GIORNO nell'intervallo [data_da, data_a] (inclusi).
    File separato da MBC/tr, come richiesto - non contiene gli altri due fogli del template."""
    from stampa_mbc import TEMPLATE_PATH  # stesso file sorgente templates_dop.xlsx

    shutil.copy(TEMPLATE_PATH, output_path)
    wb = openpyxl.load_workbook(output_path)
    for nome in list(wb.sheetnames):
        if nome != FOGLIO:
            del wb[nome]
    ws_template = wb[FOGLIO]

    n_giorni = (data_a - data_da).days + 1
    for i in range(n_giorni):
        giorno = data_da + _dt.timedelta(days=i)
        ws = ws_template if i == 0 else wb.copy_worksheet(ws_template)
        ws.title = giorno.strftime("%d-%m-%Y")
        _compila_rbc(ws, client, caseificio_id, giorno)

    wb.save(output_path)
    return output_path
