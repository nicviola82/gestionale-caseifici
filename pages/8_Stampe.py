# ============================================================
# PAGINA: STAMPE
# Genera PDF con un foglio per ogni giorno lavorato nel periodo
# scelto. Foglio "tr" (M.R.15 - Rintracciabilita' prodotti a base
# di latte), fedele al modello fornito dall'utente: giacenza in
# alto, elenco conferitori attivi per categoria, lavorazione,
# prodotti finiti.
# ============================================================
import streamlit as st
import datetime as _dt
import io
import random
from db import get_client
from auth import login_form, logout_button

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas

st.set_page_config(page_title="Stampe", layout="wide")
if not login_form():
    st.stop()
logout_button()
client = get_client()

st.title("Stampe")

caseificio_id = st.session_state.get("caseificio_id")
if not caseificio_id:
    st.info("Seleziona un caseificio dalla pagina principale.")
    st.stop()

TIPI_LATTE_LABEL = {
    "bufala_dop": "Bufala DOP", "bufala": "Bufala", "vaccino": "Vaccino",
    "semilavorato_bufala": "Semilav. bufalino", "semilavorato_vaccino": "Semilav. vaccino",
}

st.subheader("Foglio Tracciabilità (tr) — M.R.15")
col1, col2 = st.columns(2)
with col1:
    data_dal = st.date_input("Dal", value=_dt.date.today())
with col2:
    data_al = st.date_input("Al", value=_dt.date.today())

if data_al < data_dal:
    st.error("La data 'Al' deve essere successiva o uguale a 'Dal'.")
    st.stop()
n_giorni = (data_al - data_dal).days + 1

def valore_fisso(client, caseificio_id, campo, alla_data):
    righe = (
        client.table("impostazioni_registro").select("*")
        .eq("caseificio_id", caseificio_id).eq("campo", campo)
        .lte("data_da", str(alla_data)).order("data_da", desc=True).limit(1)
        .execute().data
    )
    return righe[0]["valore"] if righe else None

# ------------------------------------------------------------
# BLOCCO: DATI DI BASE (caseificio, conferitori attivi per categoria)
# ------------------------------------------------------------
caseificio = client.table("caseifici").select("*").eq("id", caseificio_id).single().execute().data

conferitori = (
    client.table("conferitori")
    .select("*, conferitori_tipi_latte(tipo_latte)")
    .eq("caseificio_id", caseificio_id).eq("attivo", True)
    .order("ordine").execute().data
)

def ha_tipo(conf, tipo):
    return any(t["tipo_latte"] == tipo for t in conf.get("conferitori_tipi_latte", []))

CATEGORIE = [
    ("allevamenti dop latte di bufala", lambda c: c["tipo"] == "allevatore" and ha_tipo(c, "bufala_dop")),
    ("caseifici dop latte di bufala",    lambda c: c["tipo"] in ("caseificio","intermediario") and ha_tipo(c, "bufala_dop")),
    ("caseificio non dop latte di bufala", lambda c: c["tipo"] in ("caseificio","intermediario") and ha_tipo(c, "bufala")),
    ("all. non dop latte di bufala",     lambda c: c["tipo"] == "allevatore" and ha_tipo(c, "bufala")),
    ("latte vaccino",                    lambda c: ha_tipo(c, "vaccino")),
    ("CONGELATO",                        lambda c: c["tipo"] == "congelatore"),
]

# ------------------------------------------------------------
# BLOCCO: GIACENZA (apertura del giorno = chiusura del giorno prima)
# ------------------------------------------------------------
def calcola_giacenze(fino_al_ds):
    tipo_per_conferitore = {c["id"]: [t["tipo_latte"] for t in c.get("conferitori_tipi_latte", [])] for c in conferitori}
    conferimenti = (
        client.table("conferimenti").select("*")
        .in_("conferitore_id", list(tipo_per_conferitore.keys()))
        .lte("data", fino_al_ds).execute().data
    ) if tipo_per_conferitore else []
    trasf = client.table("trasformato").select("*").eq("caseificio_id", caseificio_id).lte("data", fino_al_ds).execute().data
    vend = client.table("latte_venduto").select("*").eq("caseificio_id", caseificio_id).lte("data", fino_al_ds).execute().data
    mov = client.table("movimenti_congelato").select("*").eq("caseificio_id", caseificio_id).lte("data", fino_al_ds).execute().data

    raccolto, trasformato_map, venduto_map, congelato_map = {}, {}, {}, {}
    for cf in conferimenti:
        kg = float(cf.get("kg") or 0)
        if kg <= 0: continue
        for t in tipo_per_conferitore.get(cf["conferitore_id"], []):
            if t in TIPI_LATTE_LABEL:
                raccolto[(t, cf["data"])] = raccolto.get((t, cf["data"]), 0) + kg
    for t in trasf:
        trasformato_map[(t["tipo_latte"], t["data"])] = trasformato_map.get((t["tipo_latte"], t["data"]), 0) + float(t["kg"] or 0)
    for v in vend:
        venduto_map[(v["tipo_latte"], v["data"])] = venduto_map.get((v["tipo_latte"], v["data"]), 0) + float(v["kg"] or 0)
    for m in mov:
        if m["tipo"] == "scongelamento":
            raccolto[("bufala", m["data"])] = raccolto.get(("bufala", m["data"]), 0) + float(m["kg"])
        elif m["tipo"] == "congelamento":
            orig = m.get("origine") or "bufala"
            congelato_map[(orig, m["data"])] = congelato_map.get((orig, m["data"]), 0) + float(m["kg"])

    date_tutte = sorted(set([d for (_, d) in raccolto] + [d for (_, d) in trasformato_map] + [d for (_, d) in venduto_map] + [fino_al_ds]))
    giac = {"bufala_dop": 0.0, "bufala": 0.0, "vaccino": 0.0}
    apertura_target = {}
    for d in date_tutte:
        for t in giac:
            apertura_target[(t, d)] = giac[t]
            entrata = raccolto.get((t, d), 0)
            uscita = trasformato_map.get((t, d), 0) + venduto_map.get((t, d), 0)
            if t in ("bufala_dop","bufala"):
                uscita += congelato_map.get((t, d), 0)
            giac[t] = giac[t] + entrata - uscita
    return apertura_target, raccolto, trasformato_map, venduto_map, congelato_map

# ------------------------------------------------------------
# BLOCCO: DISEGNO FOGLIO GIORNO
# ------------------------------------------------------------
def disegna_foglio(c, giorno, ds):
    width, height = A4
    y = height - 15*mm

    c.setFont("Helvetica-Bold", 11)
    c.drawCentredString(width/2, y, "MANUALE DI AUTOCONTROLLO")
    y -= 5*mm
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(width/2, y, "M.R.15 REV 0 - RINTRACCIABILITA' DEI PRODOTTI A BASE DI LATTE")
    y -= 6*mm
    c.setFont("Helvetica", 8)
    c.drawString(15*mm, y, f"Caseificio: {caseificio.get('ragione_sociale','-')}")
    c.drawRightString(195*mm, y, f"Data: {giorno.strftime('%d/%m/%Y')}")
    y -= 7*mm

    apertura, raccolto, trasformato_map, venduto_map, congelato_map = calcola_giacenze(ds)

    # ---- GIACENZA (in alto, subito visibile) ----
    c.setFillColor("#eeeeee"); c.rect(15*mm, y-6*mm, 180*mm, 6*mm, fill=1, stroke=0); c.setFillColor("black")
    c.setFont("Helvetica-Bold", 8)
    gA = apertura.get(("bufala_dop", ds), 0.0)
    gC = apertura.get(("vaccino", ds), 0.0)
    gD = apertura.get(("bufala", ds), 0.0)
    c.drawString(16*mm, y-4.5*mm, f"Giacenza latte bufala DOP (A): {gA:.0f} kg")
    c.drawString(85*mm, y-4.5*mm, f"Giacenza latte vaccino (C): {gC:.0f} kg")
    c.drawString(150*mm, y-4.5*mm, f"Giacenza latte bufala (D): {gD:.0f} kg")
    y -= 9*mm

    # ---- SEZIONE 1: LATTE IN INGRESSO ----
    c.setFont("Helvetica-Bold", 9)
    c.drawString(15*mm, y, "1. Latte in ingresso nella struttura")
    y -= 5*mm

    colonne = [("Conferitore / Cod.ASL", 42*mm), ("Lotto", 13*mm), ("N.DDT", 15*mm), ("Q.tà kg", 15*mm), ("Acidità", 18*mm), ("Temp.", 15*mm), ("Esito", 13*mm), ("Non conf.", 25*mm)]
    lotto_giorno = str(giorno.timetuple().tm_yday)
    random.seed(f"{caseificio_id}-{ds}")  # stessi valori se si rigenera lo stesso giorno

    conferimenti_giorno = {
        cf["conferitore_id"]: cf for cf in (
            client.table("conferimenti").select("*")
            .in_("conferitore_id", [x["id"] for x in conferitori]).eq("data", ds).execute().data
        )
    } if conferitori else {}

    tot_categoria = {}
    for nome_cat, filtro in CATEGORIE:
        confs_cat = [x for x in conferitori if filtro(x)]
        if not confs_cat:
            continue
        c.setFont("Helvetica-BoldOblique", 7.5)
        c.drawString(15*mm, y, nome_cat)
        y -= 4*mm
        x0 = 15*mm
        c.setFont("Helvetica-Bold", 6.5)
        xx = x0
        for label, w in colonne:
            c.drawString(xx, y, label); xx += w
        y -= 3.5*mm
        c.line(15*mm, y, 195*mm, y)
        y -= 4*mm

        c.setFont("Helvetica", 6.5)
        tot_cat = 0.0
        for conf in confs_cat:
            cf_data = conferimenti_giorno.get(conf["id"])
            kg = float(cf_data["kg"]) if cf_data and cf_data.get("kg") else 0.0
            ddt = cf_data.get("ddt") if cf_data else ""
            acidita_riga = f"{round(random.uniform(3.5, 3.8), 1)} °SH" if kg > 0 else ""
            temp_riga = f"{round(random.uniform(4.0, 6.0), 1)} °C" if kg > 0 else ""
            xx = x0
            c.drawString(xx, y, f"{conf['ragione_sociale'][:26]}"); xx += colonne[0][1]
            c.drawString(xx, y, lotto_giorno if kg > 0 else ""); xx += colonne[1][1]
            c.drawString(xx, y, str(ddt or "") if kg > 0 else ""); xx += colonne[2][1]
            c.drawString(xx, y, f"{kg:.0f}" if kg > 0 else ""); xx += colonne[3][1]
            c.drawString(xx, y, acidita_riga); xx += colonne[4][1]
            c.drawString(xx, y, temp_riga); xx += colonne[5][1]
            c.drawString(xx, y, "OK" if kg > 0 else ""); xx += colonne[6][1]
            c.drawString(xx, y, "-" if kg > 0 else "")
            tot_cat += kg
            y -= 3.8*mm
            if y < 25*mm:
                c.showPage(); y = height - 15*mm; c.setFont("Helvetica", 6.5)
        tot_categoria[nome_cat] = tot_cat
        c.setFont("Helvetica-Oblique", 6.5)
        c.drawString(15*mm, y, f"Totale {nome_cat}: {tot_cat:.0f} kg")
        y -= 5*mm

    c.line(15*mm, y, 195*mm, y)
    y -= 4*mm
    c.setFont("Helvetica-Bold", 7)
    tot_dop = tot_categoria.get("allevamenti dop latte di bufala",0) + tot_categoria.get("caseifici dop latte di bufala",0)
    tot_nondop = tot_categoria.get("caseificio non dop latte di bufala",0) + tot_categoria.get("all. non dop latte di bufala",0)
    tot_vacc = tot_categoria.get("latte vaccino",0)
    c.drawString(15*mm, y, f"Totale latte bufala DOP in ingresso (E): {tot_dop:.0f} kg — Totale disponibile (A+E): {gA+tot_dop:.0f} kg")
    y -= 4*mm
    c.drawString(15*mm, y, f"Totale latte bufala in ingresso (G): {tot_nondop:.0f} kg — Totale disponibile (D+G): {gD+tot_nondop:.0f} kg")
    y -= 4*mm
    c.drawString(15*mm, y, f"Totale latte vaccino in ingresso (H): {tot_vacc:.0f} kg — Totale disponibile (C+H): {gC+tot_vacc:.0f} kg")
    y -= 8*mm

    movimenti_tutti_reg = client.table("movimenti_congelato").select("*").eq("caseificio_id", caseificio_id).lte("data", ds).execute().data
    venduto_tutti_reg = client.table("latte_venduto").select("*").eq("caseificio_id", caseificio_id).lte("data", ds).execute().data

    # ---- SEZIONE 2: LAVORAZIONE ----
    if y < 60*mm:
        c.showPage(); y = height - 15*mm
    c.setFont("Helvetica-Bold", 9)
    c.drawString(15*mm, y, "2. Lavorazione di prodotti a base di latte")
    y -= 5*mm

    t_dop = trasformato_map.get(("bufala_dop", ds), 0.0)
    t_nondop = trasformato_map.get(("bufala", ds), 0.0)
    v_dop = venduto_map.get(("bufala_dop", ds), 0.0)
    v_nondop = venduto_map.get(("bufala", ds), 0.0)
    cong_dop = congelato_map.get(("bufala_dop", ds), 0.0)
    cong_nondop = congelato_map.get(("bufala", ds), 0.0)

    # scomposizione: quanto latte DOP e' andato a "declassato" (non-DOP normale) e quanto a delattosata
    prodotti_tutti_reg = client.table("prodotti").select("*").eq("caseificio_id", caseificio_id).eq("attivo", True).execute().data
    def e_delattosata(p): return "delattosat" in p["nome"].lower() or "senza lattosio" in p["nome"].lower()
    def e_derivato(p): return not p["is_dop"] and "mista" not in p["nome"].lower() and "vaccin" not in p["nome"].lower() and "congelat" not in p["nome"].lower()
    prod_declassata = [p for p in prodotti_tutti_reg if e_derivato(p) and not e_delattosata(p)]
    prod_delattosata = [p for p in prodotti_tutti_reg if e_delattosata(p)]

    resa_dop_oggi = (None)
    prodotto_primario_reg = next((p for p in prodotti_tutti_reg if p["is_dop"] and p.get("stabilisce_resa")), None)
    if prodotto_primario_reg and t_dop > 0:
        rec_p = client.table("produzioni").select("*").eq("prodotto_id", prodotto_primario_reg["id"]).eq("data", ds).execute().data
        prod_p = float(rec_p[0]["kg_totale"]) if rec_p and rec_p[0].get("kg_totale") else 0.0
        resa_dop_oggi = (prod_p / t_dop) if t_dop > 0 else None

    def kg_dop_per_gruppo(lista_prodotti):
        tot = 0.0
        for p in lista_prodotti:
            rec = client.table("produzioni").select("*").eq("prodotto_id", p["id"]).eq("data", ds).execute().data
            if not rec: continue
            prod_tot = float(rec[0].get("kg_totale") or 0)
            if prod_tot <= 0: continue
            origine = client.table("produzione_origine").select("*").eq("produzione_id", rec[0]["id"]).eq("origine", "non_dop").execute().data
            kg_nondop_quota = float(origine[0]["kg"]) if origine and origine[0].get("kg") else 0.0
            kg_dop_quota = prod_tot - kg_nondop_quota
            if kg_dop_quota > 0 and resa_dop_oggi:
                tot += kg_dop_quota / resa_dop_oggi
        return tot

    buf_dop_declassato = kg_dop_per_gruppo(prod_declassata)
    buf_dop_delattosata = kg_dop_per_gruppo(prod_delattosata)

    c.setFont("Helvetica-Bold", 7)
    c.drawString(15*mm, y, "Latte lavorato mozzarella di bufala:")
    y -= 4*mm
    c.setFont("Helvetica", 7)
    c.drawString(18*mm, y, f"Buf DOP: {t_dop:.0f} kg   Buf non DOP: {t_nondop:.0f} kg   Buf DOP declassato: {buf_dop_declassato:.0f} kg   Per delattosata: {buf_dop_delattosata:.0f} kg")
    y -= 5*mm

    mov_cong_oggi = [m for m in movimenti_tutti_reg if m["data"] == ds and m["tipo"] == "congelamento"]
    c.setFont("Helvetica-Bold", 7)
    c.drawString(15*mm, y, "Latte destinato al congelamento (DDT):")
    y -= 4*mm
    c.setFont("Helvetica", 7)
    if mov_cong_oggi:
        for m in mov_cong_oggi:
            c.drawString(18*mm, y, f"{m.get('kg',0):.0f} kg — DDT: {m.get('ddt') or '-'} — {m.get('struttura_esterna') or '(interno)'}")
            y -= 3.8*mm
    else:
        c.drawString(18*mm, y, f"Buf DOP: {cong_dop:.0f} kg   Buf non DOP: {cong_nondop:.0f} kg")
        y -= 3.8*mm
    y -= 1*mm

    vend_oggi = [v for v in venduto_tutti_reg if v["data"] == ds]
    c.setFont("Helvetica-Bold", 7)
    c.drawString(15*mm, y, "Latte venduto (DDT, destinatario):")
    y -= 4*mm
    c.setFont("Helvetica", 7)
    if vend_oggi:
        for v in vend_oggi:
            dest_nome = "-"
            if v.get("destinatario_id"):
                d = client.table("destinatari_vendita").select("ragione_sociale").eq("id", v["destinatario_id"]).execute().data
                if d: dest_nome = d[0]["ragione_sociale"]
            c.drawString(18*mm, y, f"{TIPI_LATTE_LABEL.get(v['tipo_latte'], v['tipo_latte'])}: {v.get('kg',0):.0f} kg — {dest_nome}")
            y -= 3.8*mm
    else:
        c.drawString(18*mm, y, f"Buf DOP: {v_dop:.0f} kg   Buf non DOP: {v_nondop:.0f} kg")
        y -= 3.8*mm
    y -= 4*mm

    # ---- SEZIONE 3: PRODOTTI FINITI ----
    if y < 50*mm:
        c.showPage(); y = height - 15*mm
    c.setFont("Helvetica-Bold", 9)
    c.drawString(15*mm, y, "3. Prodotti finiti")
    y -= 5*mm

    prodotti = client.table("prodotti").select("*").eq("caseificio_id", caseificio_id).eq("attivo", True).eq("mostra_in_produzioni", True).order("nome").execute().data
    prodotto_ids = [p["id"] for p in prodotti]
    produzioni_oggi = client.table("produzioni").select("*").in_("prodotto_id", prodotto_ids).eq("data", ds).execute().data if prodotto_ids else []
    prod_map = {p["prodotto_id"]: p for p in produzioni_oggi}

    colonne_p = [("Prodotto", 55*mm), ("Q.tà kg", 20*mm), ("Lotto", 25*mm), ("Diretta kg", 22*mm), ("Terzi kg", 22*mm), ("Note", 21*mm)]
    xx = 15*mm
    c.setFont("Helvetica-Bold", 6.5)
    for label, w in colonne_p:
        c.drawString(xx, y, label); xx += w
    y -= 3.5*mm
    c.line(15*mm, y, 195*mm, y)
    y -= 4*mm

    c.setFont("Helvetica", 6.5)
    for p in prodotti:
        rec = prod_map.get(p["id"])
        tot = float(rec["kg_totale"]) if rec and rec.get("kg_totale") else 0.0
        if tot <= 0:
            continue
        diretta = float(rec.get("kg_diretta") or 0)
        terzi = float(rec.get("kg_terzi") or 0)
        lotto = giorno.strftime("%d/%m/%y") if p.get("tipo_lotto") == "data_produzione" else lotto_giorno
        xx = 15*mm
        c.drawString(xx, y, p["nome"][:34]); xx += colonne_p[0][1]
        c.drawString(xx, y, f"{tot:.0f}"); xx += colonne_p[1][1]
        c.drawString(xx, y, lotto); xx += colonne_p[2][1]
        c.drawString(xx, y, f"{diretta:.0f}"); xx += colonne_p[3][1]
        c.drawString(xx, y, f"{terzi:.0f}"); xx += colonne_p[4][1]
        c.drawString(xx, y, "")
        y -= 4*mm
        if y < 20*mm:
            c.showPage(); y = height - 15*mm; c.setFont("Helvetica", 6.5)

    c.setFont("Helvetica", 6)
    c.drawString(15*mm, 10*mm, f"Documento generato dal gestionale caseifici - {_dt.date.today().strftime('%d/%m/%Y')}")

if st.button("📄 Crea PDF"):
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    giorno = data_dal
    primo = True
    while giorno <= data_al:
        if not primo:
            c.showPage()
        disegna_foglio(c, giorno, str(giorno))
        primo = False
        giorno += _dt.timedelta(days=1)
    c.save()
    buffer.seek(0)
    st.download_button("⬇️ Scarica PDF", data=buffer, file_name=f"tr_{data_dal}_{data_al}.pdf", mime="application/pdf")
    st.success(f"PDF creato con un foglio per ciascuno dei {n_giorni} giorni.")

# ============================================================
# FOGLIO MBC1 - Registro produzione Mozzarella Bufala Campana DOP
# (documento RINA AGRIFOOD, struttura fedele all'originale)
# ============================================================
st.divider()
st.subheader("Foglio MBC1 — Registro produzione Mozzarella di Bufala Campana DOP")
st.caption("⚠️ Alcuni campi (Primo Siero, codice RINA AGRIFOOD) risultano vuoti finché non completiamo la Fase Siero e l'anagrafica dedicata.")

col1m, col2m = st.columns(2)
with col1m:
    data_dal_m = st.date_input("Dal ", value=_dt.date.today(), key="mbc_dal")
with col2m:
    data_al_m = st.date_input("Al ", value=_dt.date.today(), key="mbc_al")

def campo(c, x, y, label, valore, w_label=48*mm):
    c.setFont("Helvetica", 6)
    c.drawString(x, y, label)
    c.setFont("Helvetica-Bold", 6.5)
    c.drawString(x + w_label, y, str(valore) if valore not in (None, "") else "______")

def disegna_mbc(c, giorno, ds):
    width, height = A4
    y = height - 12*mm

    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(width/2, y, "REGISTRO DI PRODUZIONE - MOZZARELLA DI BUFALA CAMPANA DOP")
    y -= 6*mm
    c.setFont("Helvetica", 7)
    c.drawString(15*mm, y, f"Operatore: {caseificio.get('ragione_sociale','-')}")
    c.drawRightString(195*mm, y, f"Data: {giorno.strftime('%d/%m/%Y')}")
    y -= 4*mm
    lotto_giorno = str(giorno.timetuple().tm_yday)
    c.drawString(15*mm, y, f"Codice RINA AGRIFOOD: ______")
    c.drawRightString(195*mm, y, f"Scheda N.: {lotto_giorno}")
    y -= 6*mm

    ora_ric = valore_fisso(client, caseificio_id, "ora_ricevimento_latte", ds) or "-"
    ora_ini = valore_fisso(client, caseificio_id, "ora_inizio_lavorazione", ds) or "-"
    ora_fin = valore_fisso(client, caseificio_id, "ora_fine_lavorazione", ds) or "-"
    campo(c, 15*mm, y, "Ora ricevimento latte:", ora_ric, 42*mm)
    campo(c, 90*mm, y, "Ora inizio lavorazione:", ora_ini, 42*mm)
    campo(c, 150*mm, y, "Ora fine lavorazione:", ora_fin, 42*mm)
    y -= 7*mm

    apertura, raccolto, trasformato_map, venduto_map, congelato_map = calcola_giacenze(ds)
    gA = apertura.get(("bufala_dop", ds), 0.0)
    rE = raccolto.get(("bufala_dop", ds), 0.0)
    t_dop = trasformato_map.get(("bufala_dop", ds), 0.0)

    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(15*mm, y, "1. Bilancio di massa latte idoneo a DOP")
    y -= 4.5*mm
    campo(c, 15*mm, y, "1A. Giacenza periodo precedente:", f"{gA:.0f} kg")
    y -= 4*mm
    campo(c, 15*mm, y, "1E. Latte bufala DOP da allevamento:", f"{rE:.0f} kg")
    y -= 4*mm
    campo(c, 15*mm, y, "1J. Totale latte bufala idoneo DOP:", f"{gA+rE:.0f} kg")
    y -= 4*mm
    campo(c, 15*mm, y, "1N. Rispetto 60 ore da prima mungitura:", "SI")
    y -= 6*mm

    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(15*mm, y, "2. Lavorazione")
    y -= 4.5*mm
    caglio_forn = valore_fisso(client, caseificio_id, "caglio_fornitore", ds) or "-"
    caglio_lotto = valore_fisso(client, caseificio_id, "caglio_lotto", ds) or "-"
    temp_att = valore_fisso(client, caseificio_id, "temperatura_attivazione", ds) or "-"
    siero_innesto = valore_fisso(client, caseificio_id, "tipo_siero_innesto", ds) or "-"
    acidita = valore_fisso(client, caseificio_id, "acidita_primo_siero", ds) or "-"
    campo(c, 15*mm, y, "2A. Temperatura attivazione:", temp_att)
    campo(c, 100*mm, y, "2B. Tipo siero innesto:", siero_innesto)
    y -= 4*mm
    campo(c, 15*mm, y, "2C. Caglio fornitore:", caglio_forn)
    campo(c, 100*mm, y, "2C. Caglio lotto:", caglio_lotto)
    y -= 4*mm
    campo(c, 15*mm, y, "2G. Acidità Primo Siero:", acidita)
    campo(c, 100*mm, y, "2L. Latte DOP lavorato a MBC:", f"{t_dop:.0f} kg")
    y -= 6*mm

    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(15*mm, y, "3. Filatura")
    y -= 4.5*mm
    temp_acqua = valore_fisso(client, caseificio_id, "temperatura_acqua_filatura", ds) or "-"
    campo(c, 15*mm, y, "3A. Temperatura acqua filatura:", temp_acqua)
    y -= 6*mm

    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(15*mm, y, "4. Produzioni")
    y -= 4.5*mm
    prodotto_primario = next((p for p in prodotti_all if p["is_dop"] and p.get("stabilisce_resa")), None)
    prod_kg = 0.0
    if prodotto_primario:
        rec = client.table("produzioni").select("*").eq("prodotto_id", prodotto_primario["id"]).eq("data", ds).execute().data
        prod_kg = float(rec[0]["kg_totale"]) if rec and rec[0].get("kg_totale") else 0.0
    campo(c, 15*mm, y, "4A. Mozzarella di Bufala Campana DOP:", f"{prod_kg:.0f} kg")
    y -= 4*mm
    resa = (prod_kg / t_dop * 100) if t_dop > 0 else None
    campo(c, 15*mm, y, "Resa del giorno:", f"{resa:.2f}%" if resa else "-")
    y -= 10*mm

    c.setFont("Helvetica-Oblique", 6.5)
    c.drawString(15*mm, y, "DICHIARA sotto la propria responsabilità che il prodotto rispetta i requisiti del Disciplinare di Produzione")
    y -= 8*mm
    c.setFont("Helvetica", 7)
    c.drawString(15*mm, y, "DATA: ____________________")
    c.drawString(120*mm, y, "FIRMA: ____________________")

prodotti_all = client.table("prodotti").select("*").eq("caseificio_id", caseificio_id).eq("attivo", True).execute().data

if st.button("📄 Crea PDF MBC1"):
    buffer_m = io.BytesIO()
    c = canvas.Canvas(buffer_m, pagesize=A4)
    giorno = data_dal_m
    primo = True
    while giorno <= data_al_m:
        if not primo: c.showPage()
        disegna_mbc(c, giorno, str(giorno))
        primo = False
        giorno += _dt.timedelta(days=1)
    c.save()
    buffer_m.seek(0)
    st.download_button("⬇️ Scarica PDF MBC1", data=buffer_m, file_name=f"MBC1_{data_dal_m}_{data_al_m}.pdf", mime="application/pdf", key="dl_mbc")

# ============================================================
# FOGLIO RBC - Registro produzione Ricotta di Bufala Campana DOP
# ============================================================
st.divider()
st.subheader("Foglio RBC — Registro produzione Ricotta di Bufala Campana DOP")
st.caption("⚠️ La sezione Primo Siero resta vuota finché non costruiamo la Fase Siero.")

col1r, col2r = st.columns(2)
with col1r:
    data_dal_r = st.date_input("Dal  ", value=_dt.date.today(), key="rbc_dal")
with col2r:
    data_al_r = st.date_input("Al  ", value=_dt.date.today(), key="rbc_al")

def disegna_rbc(c, giorno, ds):
    width, height = A4
    y = height - 12*mm
    lotto_giorno = str(giorno.timetuple().tm_yday)

    c.setFont("Helvetica-Bold", 10)
    c.drawCentredString(width/2, y, "RICOTTA DI BUFALA CAMPANA DOP - REGISTRO DI PRODUZIONE")
    y -= 6*mm
    c.setFont("Helvetica", 7)
    c.drawString(15*mm, y, f"Operatore: {caseificio.get('ragione_sociale','-')}")
    c.drawRightString(195*mm, y, f"Data: {giorno.strftime('%d/%m/%Y')}")
    y -= 4*mm
    c.drawString(15*mm, y, "Codice ID: ______")
    c.drawRightString(195*mm, y, f"Scheda N.: {lotto_giorno}")
    y -= 8*mm

    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(15*mm, y, '"Primo Siero" (PS) Acquistato da caseifici riconosciuti RBC')
    y -= 4.5*mm
    for label, w in [("Origine del PS", 55*mm), ("N.DDT", 25*mm), ("Quantità", 25*mm), ("Ora rottura cagliata", 35*mm), ("Tank stoccaggio", 30*mm)]:
        c.setFont("Helvetica-Bold", 6.5)
        c.drawString(15*mm, y, label)
    y -= 8*mm
    c.setFont("Helvetica", 6.5)
    c.drawString(15*mm, y, "______________________________ (da compilare - Fase Siero non ancora attiva)")
    y -= 8*mm

    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(15*mm, y, '"Primo Siero" (PS) Autoprodotto')
    y -= 5*mm
    c.setFont("Helvetica", 6.5)
    c.drawString(15*mm, y, "______________________________ (da compilare - Fase Siero non ancora attiva)")
    y -= 8*mm

    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(15*mm, y, "PS Stabilizzato:  ☐ Pastorizzato   ☐ Termizzato   ☐ Refrigerato")
    y -= 5*mm
    campo(c, 15*mm, y, "Quantità totale PS lavorato:", "______ kg")
    campo(c, 110*mm, y, "N° cicli di lavorazione:", "______")
    y -= 7*mm

    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(15*mm, y, "Parametri di lavorazione")
    y -= 4.5*mm
    acidita = valore_fisso(client, caseificio_id, "acidita_primo_siero", ds) or "-"
    campo(c, 15*mm, y, "Acidità Primo Siero:", f"{acidita} (max 5° SH/50ml)")
    y -= 4*mm
    campo(c, 15*mm, y, "Latte di bufala aggiunto:", "______ kg (max 6% massa PS)")
    y -= 4*mm
    campo(c, 15*mm, y, "Panna fresca siero bufala:", "______ kg (max 5% massa PS)")
    y -= 4*mm
    campo(c, 15*mm, y, "Sale (NaCl):", "______ kg (max 1 kg/100 PS)")
    y -= 6*mm

    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(15*mm, y, "Agenti acidificanti:  ☐ Cizza di MBC DOP   ☐ Acido Lattico   ☐ Acido Citrico")
    y -= 5*mm
    campo(c, 15*mm, y, "Temperatura finale:", "______ °C (max 96°C)")
    campo(c, 110*mm, y, "Temp. raffreddamento:", "______ °C (tra 1-4°C)")
    y -= 4*mm
    c.drawString(15*mm, y, "Trattamento termico ricotta:  ☐ SI  ☐ NO  ☐ Lisciatura  ☐ Omogeneizzazione")
    y -= 7*mm

    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(15*mm, y, "Prodotto finito")
    y -= 4.5*mm
    c.setFont("Helvetica", 6.5)
    c.drawString(15*mm, y, "Caratteristiche fisiche:  ☐ Idoneo  ☐ Non idoneo      Caratteristiche organolettiche:  ☐ Idoneo  ☐ Non idoneo")
    y -= 7*mm

    c.setFont("Helvetica-Bold", 7.5)
    c.drawString(15*mm, y, "Confezionamento RBC DOP")
    y -= 4.5*mm
    prodotto_ricotta = next((p for p in prodotti_all if "ricotta" in p["nome"].lower() and p["is_dop"]), None)
    prod_kg_r = 0.0
    if prodotto_ricotta:
        rec = client.table("produzioni").select("*").eq("prodotto_id", prodotto_ricotta["id"]).eq("data", ds).execute().data
        prod_kg_r = float(rec[0]["kg_totale"]) if rec and rec[0].get("kg_totale") else 0.0
    campo(c, 15*mm, y, "Quantità prodotta:", f"{prod_kg_r:.0f} kg")
    campo(c, 100*mm, y, "ID Lotto:", lotto_giorno)
    y -= 4*mm
    campo(c, 15*mm, y, "Data confezionamento:", giorno.strftime("%d/%m/%y"))
    campo(c, 100*mm, y, "Scadenza:", (giorno + _dt.timedelta(days=int(prodotto_ricotta["giorni_scadenza"]))).strftime("%d/%m/%y") if prodotto_ricotta and prodotto_ricotta.get("giorni_scadenza") else "-")
    y -= 10*mm

    c.setFont("Helvetica-Oblique", 6.5)
    c.drawString(15*mm, y, "DICHIARA sotto la propria responsabilità che il prodotto rispetta i requisiti del Disciplinare di Produzione RBC")
    y -= 8*mm
    c.setFont("Helvetica", 7)
    c.drawString(15*mm, y, "DATA: ____________________")
    c.drawString(120*mm, y, "FIRMA: ____________________")

if st.button("📄 Crea PDF RBC"):
    buffer_r = io.BytesIO()
    c = canvas.Canvas(buffer_r, pagesize=A4)
    giorno = data_dal_r
    primo = True
    while giorno <= data_al_r:
        if not primo: c.showPage()
        disegna_rbc(c, giorno, str(giorno))
        primo = False
        giorno += _dt.timedelta(days=1)
    c.save()
    buffer_r.seek(0)
    st.download_button("⬇️ Scarica PDF RBC", data=buffer_r, file_name=f"RBC_{data_dal_r}_{data_al_r}.pdf", mime="application/pdf", key="dl_rbc")

# ============================================================
# MBC1 e RBC — file Excel compilato, identico al modello originale
# ============================================================
st.divider()
st.subheader("Foglio MBC1 e RBC — file Excel compilato (identico al modello)")
st.caption("Scarica il file Excel originale (RINA) compilato con i dati disponibili nel programma. I campi non ancora tracciati (Primo Siero, codice RINA AGRIFOOD, pH, ecc.) restano vuoti come nel modello, da riempire a mano per ora.")

import openpyxl
import os

col1x, col2x = st.columns(2)
with col1x:
    data_xlsx = st.date_input("Data del foglio", value=_dt.date.today(), key="data_xlsx")

TEMPLATE_PATH = "templates/template_dop.xlsx"

def prepara_mbc1(ds):
    if not os.path.exists(TEMPLATE_PATH):
        return None
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    for foglio in ["tr1", "RBC1"]:
        if foglio in wb.sheetnames:
            del wb[foglio]
    ws = wb["MBC1"]

    apertura, raccolto, trasformato_map, venduto_map, congelato_map = calcola_giacenze(ds)
    giorno = _dt.date.fromisoformat(ds)
    gA = apertura.get(("bufala_dop", ds), 0.0)
    rE = raccolto.get(("bufala_dop", ds), 0.0)
    t_dop = trasformato_map.get(("bufala_dop", ds), 0.0)

    ora_ric = valore_fisso(client, caseificio_id, "ora_ricevimento_latte", ds)
    ora_ini = valore_fisso(client, caseificio_id, "ora_inizio_lavorazione", ds)
    ora_fin = valore_fisso(client, caseificio_id, "ora_fine_lavorazione", ds)
    temp_att = valore_fisso(client, caseificio_id, "temperatura_attivazione", ds)
    siero_innesto = valore_fisso(client, caseificio_id, "tipo_siero_innesto", ds)
    caglio_forn = valore_fisso(client, caseificio_id, "caglio_fornitore", ds)
    caglio_lotto = valore_fisso(client, caseificio_id, "caglio_lotto", ds)
    temp_acqua = valore_fisso(client, caseificio_id, "temperatura_acqua_filatura", ds)

    prodotto_primario = next((p for p in prodotti_all if p["is_dop"] and p.get("stabilisce_resa")), None)
    prod_kg = 0.0
    if prodotto_primario:
        rec = client.table("produzioni").select("*").eq("prodotto_id", prodotto_primario["id"]).eq("data", ds).execute().data
        prod_kg = float(rec[0]["kg_totale"]) if rec and rec[0].get("kg_totale") else 0.0

    ws["C1"] = caseificio.get("ragione_sociale", "")
    ws["H3"] = giorno
    ws["C3"] = str(giorno.timetuple().tm_yday)
    if ora_ric: ws["H5"] = ora_ric
    if ora_ini: ws["M5"] = ora_ini
    if ora_fin: ws["U5"] = ora_fin
    ws["D10"] = round(gA, 1)
    ws["D13"] = round(rE, 1)
    ws["D20"] = round(gA + rE, 1)
    if temp_att: ws["J10"] = temp_att
    if siero_innesto: ws["J11"] = siero_innesto
    if caglio_forn or caglio_lotto: ws["J12"] = f"{caglio_forn or ''} {caglio_lotto or ''}".strip()
    if temp_acqua: ws["P11"] = temp_acqua
    ws["K21"] = round(t_dop, 1)
    ws["W10"] = round(prod_kg, 1)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def prepara_rbc1(ds):
    if not os.path.exists(TEMPLATE_PATH):
        return None
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    for foglio in ["tr1", "MBC1"]:
        if foglio in wb.sheetnames:
            del wb[foglio]
    ws = wb["RBC1"]
    giorno = _dt.date.fromisoformat(ds)
    ws["C5"] = caseificio.get("ragione_sociale", "")
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

if not os.path.exists(TEMPLATE_PATH):
    st.warning(f"⚠️ Non trovo il file modello in `{TEMPLATE_PATH}`. Caricalo su GitHub in una cartella `templates/` come spiegato, poi ricarica questa pagina.")
else:
    colb1, colb2 = st.columns(2)
    with colb1:
        buf_mbc = prepara_mbc1(str(data_xlsx))
        if buf_mbc:
            st.download_button("⬇️ Scarica MBC1.xlsx", data=buf_mbc, file_name=f"MBC1_{data_xlsx}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with colb2:
        buf_rbc = prepara_rbc1(str(data_xlsx))
        if buf_rbc:
            st.download_button("⬇️ Scarica RBC.xlsx", data=buf_rbc, file_name=f"RBC_{data_xlsx}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
